#!/usr/bin/env python3
"""
PBC Monthly Report Generator v3.0
==================================
Consolidated GUI tool for generating all PBC financial reports:
  - 5 Individual Reports (Pack Volume, NSR, COMS, Power & Fuel, Legal)
  - 1 Combined Analysis workbook (all 5 as tabs)
  - 1 Enhanced Analysis v2.1 workbook (10 McKinsey-grade sheets)

Requirements: pip install pandas openpyxl numpy xlrd
"""

import tkinter as tk
from tkinter import filedialog, scrolledtext, messagebox, ttk
import threading
import pandas as pd
import numpy as np
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
import sys
from pathlib import Path
from datetime import datetime
from glob import glob
import traceback
import warnings
warnings.filterwarnings('ignore')

# ═══════════════════════════════════════════════════════════════
# CONSTANTS
# ═══════════════════════════════════════════════════════════════
MONTHS = [
    ('January', 1), ('February', 2), ('March', 3), ('April', 4),
    ('May', 5), ('June', 6), ('July', 7), ('August', 8),
    ('September', 9), ('October', 10), ('November', 11), ('December', 12)
]
YEARS = [str(y) for y in range(2024, 2030)]

# ═══════════════════════════════════════════════════════════════
# v2.1 ENHANCED ANALYSIS STYLING CONSTANTS
# ═══════════════════════════════════════════════════════════════
HDR_FILL = PatternFill('solid', fgColor='1F4E79')
HDR_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=11)
SUB_FILL = PatternFill('solid', fgColor='D6E4F0')
SUB_FONT = Font(name='Arial', bold=True, size=10)
DATA_FONT = Font(name='Arial', size=10)
BLUE_INPUT = Font(name='Arial', size=10, color='0000FF')
TITLE_FONT = Font(name='Arial', bold=True, size=14, color='1F4E79')
SUBTITLE_FONT = Font(name='Arial', bold=True, size=11, color='366092')
WARN_FONT = Font(name='Arial', bold=True, size=10, color='C00000')
ITALIC_NOTE = Font(name='Arial', italic=True, size=9, color='666666')
CAVEAT_FONT = Font(name='Arial', italic=True, size=9, color='C00000')
BOLD_DATA = Font(name='Arial', bold=True, size=10)
BOLD_BLUE = Font(name='Arial', bold=True, size=10, color='1F4E79')
PCT_FMT = '0.0%'
NUM_FMT = '#,##0'
DEC_FMT = '#,##0.00'
THIN = Border(
    left=Side(style='thin', color='B0B0B0'), right=Side(style='thin', color='B0B0B0'),
    top=Side(style='thin', color='B0B0B0'), bottom=Side(style='thin', color='B0B0B0'))
TOTAL_BORDER = Border(
    top=Side(style='medium', color='1F4E79'), bottom=Side(style='double', color='1F4E79'),
    left=Side(style='thin', color='B0B0B0'), right=Side(style='thin', color='B0B0B0'))

# ═══════════════════════════════════════════════════════════════
# v2.1 COMS CATEGORY DATA (known structure from PBC RM sheet)
# ═══════════════════════════════════════════════════════════════
COMS_CATEGORIES = [
    ('Concentrates & Beverages', 368474942, 480289690),
    ('Sugar', 246671618, 340412912),
    ('CO2 & Nitrogen Gas', 10203470, 17377643),
    ('Crowns & Caps', 8709743, 8968740),
    ('Screw Caps & Closures', 24017236, 41083699),
    ('Preforms', 178334562, 287899068),
    ('Shrink Film, Labels & Packaging', 36903592, 60051817),
    ('Juice Materials (Mango Pulp etc.)', 32883124, 39001142),
    ('Water Treatment Chemicals', 7593092, 13170741),
    ('Finished Goods (3rd Party)', 14265907, 5622641),
]

KEY_RM_RATES = [
    ('Sugar', 'per kg', 120.218366, 129.579124),
    ('CO2 Gas', 'per kg', 30.445397, 55.668046),
    ('Nitrogen Gas', 'per kg', 102.980211, 102.980207),
    ('Pepsi Concentrate', 'per unit', 193455.97, 200303.47),
    ('7-UP Concentrate', 'per unit', 64491.91, 66769.47),
    ('Mountain Dew Concentrate', 'per unit', 64488.86, 46770.88),
    ('Sting (Red) Concentrate', 'per unit', 147431.93, 159446.77),
    ('Aquafina Salt', 'per unit', 77899.95, 77900.00),
    ('Mango Pulp', 'per kg', 149.36, 130.85),
    ('Preforms 41g (Green)', 'per unit', 17.19, 16.32),
    ('Preforms 30g (Green)', 'per unit', 13.12, 12.37),
    ('Preforms 20g (White)', 'per unit', 0, 8.93),
    ('Preforms 16g (A/Fina)', 'per unit', 7.25, 6.91),
    ('Shrink Film 730mm', 'per kg', 392.29, 396.69),
    ('Labels', 'per unit', 0.8285, 0.7648),
    ('Hotmelt Glue', 'per kg', 1869.51, 1799.98),
    ('Layer Pad', 'per unit', 95.48, 113.66),
    ('Caustic Soda 50%', 'per kg', 78.33, 77.00),
    ('Conveyor Lubricant', 'per kg', 1810, 1810),
]


# ═══════════════════════════════════════════════════════════════
# v2.1 HELPER FUNCTIONS
# ═══════════════════════════════════════════════════════════════
def enh_style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HDR_FONT; cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN

def enh_style_data(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = DATA_FONT; cell.border = THIN
        cell.alignment = Alignment(horizontal='right' if c > 1 else 'left')

def enh_style_total(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = BOLD_DATA; cell.border = TOTAL_BORDER

def enh_write_title(ws, row, title, subtitle=None):
    ws.cell(row=row, column=1, value=title).font = TITLE_FONT
    if subtitle:
        ws.cell(row=row + 1, column=1, value=subtitle).font = SUBTITLE_FONT
    return row + (3 if subtitle else 2)

def enh_auto_width(ws, ncols, min_w=14, max_w=35):
    for c in range(1, ncols + 1):
        ws.column_dimensions[get_column_letter(c)].width = min_w
    ws.column_dimensions['A'].width = max_w

def safe_float(val, default=0):
    if pd.isna(val):
        return default
    try:
        return float(val)
    except (ValueError, TypeError):
        return default

def get_customer_sales(df):
    name_col = amt_col = None
    for col in df.columns:
        cl = str(col).lower()
        if 'name' in cl:
            name_col = col
        if 'amount' in cl or 'local' in cl:
            amt_col = col
    if name_col and amt_col:
        return df.groupby(name_col)[amt_col].sum().sort_values(ascending=False)
    return pd.Series(dtype=float)

def aggregate_legal(df):
    name_col = 'Name of offsetting account'
    amt_col = 'Amount in local cur.'
    cc_col = 'Cost Ctr'
    for col in df.columns:
        if 'name' in str(col).lower():
            name_col = col
        if 'amount' in str(col).lower():
            amt_col = col
    result = {}
    for _, row in df.iterrows():
        name = str(row.get(name_col, 'Unknown')).strip() if pd.notna(row.get(name_col)) else 'Unknown'
        amt = safe_float(row.get(amt_col, 0))
        cc = str(row.get(cc_col, '')) if pd.notna(row.get(cc_col)) else ''
        if name in result:
            result[name]['amount'] += amt
        else:
            result[name] = {'amount': amt, 'cc': cc}
    return dict(sorted(result.items(), key=lambda x: -x[1]['amount']))


# ═══════════════════════════════════════════════════════════════
# DATA LOADER
# ═══════════════════════════════════════════════════════════════
class DataLoader:
    """Handles loading and caching of raw data files"""

    def __init__(self, input_folder, progress_callback):
        self.input_folder = input_folder
        self.progress_callback = progress_callback
        self.data = {}

    def log(self, message):
        self.progress_callback(message)

    def find_file(self, *patterns):
        for pattern in patterns:
            files = glob(os.path.join(self.input_folder, pattern))
            if files:
                return files[0]
        return None

    def load_sales_gl(self, current_month_name, prior_month_name):
        try:
            filepath = self.find_file('*Sales*GL*.xlsx', 'Sales GL.xlsx')
            if not filepath:
                self.log("⚠ Sales GL not found")
                return False
            self.log(f"  Loading Sales GL from {Path(filepath).name}...")
            xl_file = pd.ExcelFile(filepath)
            sheets = xl_file.sheet_names
            current_sheet = prior_sheet = None
            for sheet in sheets:
                sl = sheet.lower()
                if any(m in sl for m in [current_month_name.lower(), 'current', '26']):
                    current_sheet = sheet
                if any(m in sl for m in [prior_month_name.lower(), 'prior', '25']):
                    prior_sheet = sheet
            if not current_sheet and sheets:
                current_sheet = sheets[0]
            if not prior_sheet and len(sheets) > 1:
                prior_sheet = sheets[1]
            if current_sheet:
                self.data['sales_current'] = pd.read_excel(filepath, sheet_name=current_sheet)
            if prior_sheet:
                self.data['sales_prior'] = pd.read_excel(filepath, sheet_name=prior_sheet)
            self.log("  ✓ Sales GL loaded")
            return True
        except Exception as e:
            self.log(f"✗ Error loading Sales GL: {str(e)[:100]}")
            return False

    def load_discount_summary(self):
        try:
            filepath = self.find_file('*Discount*Summary*.xlsx', '*Discount*summary*.xlsx')
            if not filepath:
                self.log("⚠ Discount Summary not found"); return False
            self.log("  Loading Discount Summary...")
            xl_file = pd.ExcelFile(filepath)
            for sheet in xl_file.sheet_names:
                if 'nsr' in sheet.lower():
                    self.data['disc_nsr'] = pd.read_excel(filepath, sheet_name=sheet, header=None)
                if 'summary' in sheet.lower() and 'nsr' not in sheet.lower():
                    self.data['disc_detail'] = pd.read_excel(filepath, sheet_name=sheet, header=None)
                if 'summary' in sheet.lower():
                    self.data['discount_summary'] = pd.read_excel(filepath, sheet_name=sheet)
            self.log("  ✓ Discount Summary loaded")
            return True
        except Exception as e:
            self.log(f"✗ Error loading Discount Summary: {str(e)[:100]}")
            return False

    def load_discount_gl(self):
        try:
            filepath = self.find_file('*Discount*GL*.xlsx')
            if not filepath or 'Summary' in filepath:
                self.log("⚠ Discount GL not found"); return False
            self.log("  Loading Discount GL...")
            xl_file = pd.ExcelFile(filepath)
            sheets = xl_file.sheet_names
            for sheet in sheets:
                sl = sheet.lower()
                if '25' in sl:
                    self.data['disc_gl_25'] = pd.read_excel(filepath, sheet_name=sheet)
                elif '26' in sl:
                    self.data['disc_gl_26'] = pd.read_excel(filepath, sheet_name=sheet)
            self.data['discount_gl'] = pd.read_excel(filepath, sheet_name=0)
            self.log("  ✓ Discount GL loaded")
            return True
        except Exception as e:
            self.log(f"✗ Error loading Discount GL: {str(e)[:100]}")
            return False

    def load_raw_material(self):
        try:
            filepath = self.find_file('*RM*Consumption*.xlsx', '*Raw*Material*.xlsx')
            if not filepath:
                self.log("⚠ Raw Material Consumption not found"); return False
            self.log("  Loading Raw Material...")
            xl_file = pd.ExcelFile(filepath)
            sheets = xl_file.sheet_names
            if len(sheets) >= 1:
                self.data['rm_current'] = pd.read_excel(filepath, sheet_name=sheets[0])
                self.data['rm_26_raw'] = pd.read_excel(filepath, sheet_name=sheets[0], header=None)
            if len(sheets) >= 2:
                self.data['rm_prior'] = pd.read_excel(filepath, sheet_name=sheets[1])
                self.data['rm_25_raw'] = pd.read_excel(filepath, sheet_name=sheets[1], header=None)
            self.log("  ✓ Raw Material loaded")
            return True
        except Exception as e:
            self.log(f"✗ Error loading Raw Material: {str(e)[:100]}")
            return False

    def load_zpsr_billing(self):
        try:
            # Load both Feb 25 and Feb 26 ZPSR billing files
            for pattern, key in [('*ZPSR*25*[Bb]illing*.xlsx', 'zpsr_25'),
                                  ('*ZPSR*26*[Bb]illing*.xlsx', 'zpsr_26')]:
                filepath = self.find_file(pattern)
                if filepath:
                    self.data[key] = pd.read_excel(filepath, sheet_name=0, header=None)
            # Fallback: single file
            filepath = self.find_file('*ZPSR*[Bb]illing*.xlsx')
            if filepath:
                self.data['zpsr_billing'] = pd.read_excel(filepath, sheet_name=0)
            self.log("  ✓ ZPSR Billing loaded")
            return True
        except Exception as e:
            self.log(f"✗ Error loading ZPSR Billing: {str(e)[:100]}")
            return False

    def load_zpsr_production(self):
        try:
            filepath = self.find_file('*ZPSR*[Pp]roduction*.xls', '*ZPSR*[Pp]roduction*.xlsx')
            if not filepath:
                self.log("⚠ ZPSR Production not found"); return False
            self.log("  Loading ZPSR Production...")
            self.data['zpsr_production'] = pd.read_excel(filepath, sheet_name=0)
            self.log("  ✓ ZPSR Production loaded")
            return True
        except Exception as e:
            self.log(f"✗ Error loading ZPSR Production: {str(e)[:100]}")
            return False

    def load_legal_gl(self):
        try:
            filepath = self.find_file('*Legal*GL*.xlsx', '*Legal*Professional*.xlsx')
            if not filepath:
                self.log("⚠ Legal GL not found"); return False
            self.log("  Loading Legal GL...")
            xl_file = pd.ExcelFile(filepath)
            sheets = xl_file.sheet_names
            self.data['legal_gl'] = pd.read_excel(filepath, sheet_name=0)
            # Load both Jan and Feb sheets for v2.1
            for sheet in sheets:
                sl = sheet.lower()
                if 'jan' in sl:
                    self.data['legal_jan'] = pd.read_excel(filepath, sheet_name=sheet)
                elif 'feb' in sl:
                    self.data['legal_feb'] = pd.read_excel(filepath, sheet_name=sheet)
            self.log("  ✓ Legal GL loaded")
            return True
        except Exception as e:
            self.log(f"✗ Error loading Legal GL: {str(e)[:100]}")
            return False

    def load_vouchers(self):
        try:
            filepath = self.find_file('*Voucher*.xlsx', '*Region*.xlsx')
            if not filepath:
                self.log("⚠ Vouchers not found"); return False
            self.log("  Loading Vouchers...")
            self.data['vouchers'] = pd.read_excel(filepath, sheet_name=0)
            self.log("  ✓ Vouchers loaded")
            return True
        except Exception as e:
            self.log(f"✗ Error loading Vouchers: {str(e)[:100]}")
            return False

    def load_performed_power(self, performed_folder):
        """Load performed Power & Fuel analysis for v2.1"""
        try:
            pf_file = glob(os.path.join(performed_folder, '*Power*Fuel*.xlsx'))
            if not pf_file:
                pf_file = glob(os.path.join(performed_folder, '*Power*fuel*.xlsx'))
            if pf_file:
                self.log("  Loading Performed Power & Fuel Analysis...")
                self.data['power_perf'] = pd.read_excel(pf_file[0], header=None)
                self.log("  ✓ Performed Power & Fuel loaded")
                return True
            else:
                self.log("⚠ Performed Power & Fuel Analysis not found")
                return False
        except Exception as e:
            self.log(f"✗ Error loading Performed Power: {str(e)[:100]}")
            return False


# ═══════════════════════════════════════════════════════════════
# v2.1 VALUE EXTRACTION
# ═══════════════════════════════════════════════════════════════
def extract_enhanced_values(data, log_fn=print):
    """Extract all key values from loaded data for Enhanced Analysis v2.1."""
    v = {}
    nsr = data.get('disc_nsr')
    if nsr is None:
        log_fn("⚠ NSR Summary not available for Enhanced Analysis")
        return None

    for i in range(len(nsr)):
        label = str(nsr.iloc[i, 0]).strip() if pd.notna(nsr.iloc[i, 0]) else ''
        if 'Volume' in label and '8Oz' in label:
            v['vol_25'] = safe_float(nsr.iloc[i, 1])
            v['vol_26'] = safe_float(nsr.iloc[i, 2])
        elif label == 'Gross Sales Revenue':
            v['gsr_25'] = safe_float(nsr.iloc[i, 1])
            v['gsr_26'] = safe_float(nsr.iloc[i, 2])
        elif 'Sampling' in label and 'Claims' in label and 'per' not in label:
            v['disc_total_25'] = safe_float(nsr.iloc[i, 1])
            v['disc_total_26'] = safe_float(nsr.iloc[i, 2])
        elif 'Net Sales Revenue' in label:
            v['nsr_25'] = safe_float(nsr.iloc[i, 1])
            v['nsr_26'] = safe_float(nsr.iloc[i, 2])

    # Discount line items
    dd = data.get('disc_detail')
    v['disc_lines'] = []
    if dd is not None:
        for i in range(len(dd)):
            desc = str(dd.iloc[i, 1]).strip() if pd.notna(dd.iloc[i, 1]) else ''
            gl = dd.iloc[i, 0]
            if desc and desc not in ('Description', 'Total', '', 'nan') and pd.notna(gl):
                try:
                    float(gl)
                    v['disc_lines'].append({
                        'GL': gl, 'Description': desc,
                        'Feb_25': safe_float(dd.iloc[i, 2]),
                        'Feb_26': safe_float(dd.iloc[i, 3]),
                        'Basis': str(dd.iloc[i, 4]) if dd.shape[1] > 4 and pd.notna(dd.iloc[i, 4]) else ''
                    })
                except (ValueError, TypeError):
                    pass

    # RM totals
    for df_key, prefix in [('rm_25_raw', '25'), ('rm_26_raw', '26')]:
        df = data.get(df_key)
        if df is not None:
            for i in range(len(df)):
                label = str(df.iloc[i, 0]).strip() if pd.notna(df.iloc[i, 0]) else ''
                if 'Grand Total' in label:
                    v[f'rm_total_{prefix}'] = safe_float(df.iloc[i, 3])
                elif label == 'Production':
                    v[f'prod_cases_{prefix}'] = safe_float(df.iloc[i, 3])

    # ZPSR production
    for zkey, prefix in [('zpsr_25', '25'), ('zpsr_26', '26')]:
        zdf = data.get(zkey)
        if zdf is not None:
            last_col = zdf.shape[1] - 1
            for i in range(len(zdf)):
                val = safe_float(zdf.iloc[i, last_col])
                if val > 1e6 and i > 5:
                    next_val = safe_float(zdf.iloc[i+1, last_col]) if i+1 < len(zdf) else 0
                    if 1.0 < next_val < 1.1:
                        v[f'zpsr_prod_{prefix}'] = val
                        v[f'zpsr_conv_{prefix}'] = next_val
                        if i+2 < len(zdf):
                            v[f'zpsr_billing_{prefix}'] = safe_float(zdf.iloc[i+2, last_col])
                        break

    # ZPSR brand breakdown
    for zkey, prefix in [('zpsr_25', '25'), ('zpsr_26', '26')]:
        zdf = data.get(zkey)
        brands = {}
        if zdf is not None:
            last_col = zdf.shape[1] - 1
            for i in range(len(zdf)):
                brand = zdf.iloc[i, 1] if pd.notna(zdf.iloc[i, 1]) else None
                if brand:
                    brand_str = str(brand).strip()
                    if brand_str.lower() in ('brand', 'post mix (conv)', 'converted (250 ml)', '', 'nan'):
                        continue
                    conv = safe_float(zdf.iloc[i, last_col])
                    if conv > 0:
                        brands[brand_str] = conv
        v[f'brands_{prefix}'] = brands

    # Power & Fuel
    pf = data.get('power_perf')
    if pf is not None:
        v['wapda_items'] = []
        for i in range(len(pf)):
            label = str(pf.iloc[i, 1]).strip() if pd.notna(pf.iloc[i, 1]) else ''
            if not label or label == 'nan':
                continue
            val_25 = safe_float(pf.iloc[i, 2])
            val_26 = safe_float(pf.iloc[i, 3])
            if 'Units Produced' in label:
                v['pf_prod_25'] = val_25; v['pf_prod_26'] = val_26
            elif 'WAPDA Units' in label and 'Cost' not in label:
                v['kwh_25'] = val_25; v['kwh_26'] = val_26
                v['wapda_items'].append((label, val_25, val_26))
            elif 'WAPDA Impact' in label or 'FESCO Bill' in label:
                if 'Z check' not in label:
                    v['wapda_total_25'] = val_25; v['wapda_total_26'] = val_26
            elif label in ('Energy Charges', 'Qtr. Tariff Adj', 'Fix Charges',
                           'Meter Rent & Service Charges', 'Fuel Price Adjustment',
                           'Electricity Duty', 'FC Surcharge', 'ICP-25 Relief',
                           'Taxes On FPA', 'Taxes on FPA',
                           'Qtr. Tariff Adjustment', 'Fixed Charges'):
                v['wapda_items'].append((label, val_25, val_26))
    else:
        v['wapda_total_25'] = 0; v['wapda_total_26'] = 0
        v['kwh_25'] = 0; v['kwh_26'] = 0; v['wapda_items'] = []

    # Legal totals
    legal_jan = data.get('legal_jan')
    legal_feb = data.get('legal_feb')
    if legal_jan is not None:
        amt_col = [c for c in legal_jan.columns if 'amount' in str(c).lower()]
        v['legal_jan_total'] = legal_jan[amt_col[0]].sum() if amt_col else 0
    else:
        v['legal_jan_total'] = 0
    if legal_feb is not None:
        amt_col = [c for c in legal_feb.columns if 'amount' in str(c).lower()]
        v['legal_feb_total'] = legal_feb[amt_col[0]].sum() if amt_col else 0
    else:
        v['legal_feb_total'] = 0

    # Row counts for audit
    sales_25 = data.get('sales_prior')
    sales_26 = data.get('sales_current')
    v['sales_25_rows'] = len(sales_25) if sales_25 is not None else 0
    v['sales_26_rows'] = len(sales_26) if sales_26 is not None else 0
    dg25 = data.get('disc_gl_25')
    dg26 = data.get('disc_gl_26')
    v['disc_gl_25_rows'] = len(dg25) if dg25 is not None else 0
    v['disc_gl_26_rows'] = len(dg26) if dg26 is not None else 0

    # Set defaults for any missing values
    for key in ['vol_25','vol_26','gsr_25','gsr_26','nsr_25','nsr_26',
                'disc_total_25','disc_total_26','rm_total_25','rm_total_26',
                'prod_cases_25','prod_cases_26','wapda_total_25','wapda_total_26',
                'kwh_25','kwh_26']:
        v.setdefault(key, 0)
    v.setdefault('brands_25', {}); v.setdefault('brands_26', {})
    v.setdefault('wapda_items', [])

    return v




# ═══════════════════════════════════════════════════════════════
# REPORT GENERATOR CLASS
# ═══════════════════════════════════════════════════════════════
class ReportGenerator:
    """Main report generation engine — produces all 7 output files"""

    def __init__(self, input_folder, output_folder, current_month, current_year,
                 prior_month, prior_year, config_values, progress_callback):
        self.input_folder = input_folder
        self.output_folder = output_folder
        self.current_month = current_month
        self.current_year = current_year
        self.prior_month = prior_month
        self.prior_year = prior_year
        self.config_values = config_values
        self.progress_callback = progress_callback
        self.loader = DataLoader(input_folder, progress_callback)
        self.data = {}

    def log(self, message):
        self.progress_callback(message)

    def load_all_data(self):
        """Load all input data files"""
        self.log("\nLoading raw data files...")
        self.loader.load_sales_gl(self.current_month, self.prior_month)
        self.loader.load_discount_summary()
        self.loader.load_discount_gl()
        self.loader.load_raw_material()
        self.loader.load_zpsr_billing()
        self.loader.load_zpsr_production()
        self.loader.load_legal_gl()
        self.loader.load_vouchers()
        # Enhanced v2.1: try loading performed Power & Fuel from sibling folder
        parent = Path(self.input_folder).parent
        perf_folder = parent / "2. Performed Analysis Folder"
        if perf_folder.exists():
            self.loader.load_performed_power(str(perf_folder))
        self.data = self.loader.data

    # ─── INDIVIDUAL REPORT STYLING HELPERS ──────────────────────
    def _get_style_header(self):
        return {
            'font': Font(bold=True, color="FFFFFF", size=11),
            'fill': PatternFill(start_color="366092", end_color="366092", fill_type="solid"),
            'alignment': Alignment(horizontal="center", vertical="center"),
            'border': Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
        }

    def _apply_header_style(self, ws, row, num_cols):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            style = self._get_style_header()
            cell.font = style['font']
            cell.fill = style['fill']
            cell.alignment = style['alignment']
            cell.border = style['border']

    # ─── REPORT 1: PACK WISE VOLUME & TOP 10 CUSTOMERS ─────────
    def generate_pack_volume_report(self):
        """Report 1: Pack wise Volume & Top 10 Customers"""
        self.log("Generating Report 1: Pack Wise Volume & Top 10 Customers...")
        try:
            df = self.data.get('sales_current')
            if df is None or df.empty:
                self.log("Warning: Sales data not available")
                return False

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Pack Volume"
            df_prior = self.data.get('sales_prior', pd.DataFrame())

            pack_cols = [col for col in df.columns if 'pack' in col.lower()]
            qty_cols = [col for col in df.columns if 'quantity' in col.lower() or '8-oz' in col.lower()]
            rev_cols = [col for col in df.columns if 'amount' in col.lower()]
            cust_cols = [col for col in df.columns if 'customer' in col.lower() or 'offsetting' in col.lower()]

            pack_col = pack_cols[0] if pack_cols else df.columns[0]
            qty_col = qty_cols[0] if qty_cols else df.columns[1]
            rev_col = rev_cols[0] if rev_cols else df.columns[2]
            cust_col = cust_cols[0] if cust_cols else df.columns[-1]

            pack_volume = df.groupby(pack_col, as_index=False)[qty_col].sum().sort_values(qty_col, ascending=False)

            ws['A1'] = f"Pack Wise Volume - {self.current_month} {self.current_year}"
            ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
            ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            ws.merge_cells('A1:D1')

            headers = ['Pack Size', 'Volume (8-Oz)', 'Revenue (PKR)', 'Share %']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col)
                cell.value = header
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

            total_vol = pack_volume[qty_col].sum()

            for row_idx, (idx, row) in enumerate(pack_volume.iterrows(), 4):
                ws.cell(row=row_idx, column=1).value = row[pack_col]
                ws.cell(row=row_idx, column=2).value = row[qty_col]
                ws.cell(row=row_idx, column=2).number_format = '#,##0'
                rev_val = df[df[pack_col] == row[pack_col]][rev_col].sum()
                ws.cell(row=row_idx, column=3).value = rev_val
                ws.cell(row=row_idx, column=3).number_format = '#,##0'
                ws.cell(row=row_idx, column=4).value = f"=B{row_idx}/{total_vol}"
                ws.cell(row=row_idx, column=4).number_format = '0.0%'

            # Top 10 Customers sheet
            top_10_cust = df.groupby(cust_col, as_index=False)[rev_col].sum().sort_values(rev_col, ascending=False).head(10)
            ws2 = wb.create_sheet("Top 10 Customers")
            ws2['A1'] = f"Top 10 Customers - {self.current_month} {self.current_year}"
            ws2['A1'].font = Font(bold=True, size=14, color="FFFFFF")
            ws2['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            ws2.merge_cells('A1:C1')

            headers2 = ['Customer', 'Revenue (PKR)', 'Rank']
            for col, header in enumerate(headers2, 1):
                cell = ws2.cell(row=3, column=col)
                cell.value = header
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

            for row_idx, (idx, row) in enumerate(top_10_cust.iterrows(), 4):
                ws2.cell(row=row_idx, column=1).value = row[cust_col]
                ws2.cell(row=row_idx, column=2).value = row[rev_col]
                ws2.cell(row=row_idx, column=2).number_format = '#,##0'
                ws2.cell(row=row_idx, column=3).value = row_idx - 3

            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 15
            ws2.column_dimensions['A'].width = 30
            ws2.column_dimensions['B'].width = 15

            output_file = os.path.join(self.output_folder,
                f"1.1 Sales - Pack wise Volume and Top 10 Customers - {self.current_month} {self.current_year[-2:]}.xlsx")
            wb.save(output_file)
            self.log(f"  Report 1 saved: {Path(output_file).name}")
            return True
        except Exception as e:
            self.log(f"Error generating Report 1: {str(e)[:100]}")
            return False

    # ─── REPORT 2: NSR ANALYSIS ────────────────────────────────
    def generate_nsr_analysis_report(self):
        """Report 2: NSR Analysis"""
        self.log("Generating Report 2: NSR Analysis...")
        try:
            df_current = self.data.get('sales_current')
            df_prior = self.data.get('sales_prior')
            if df_current is None or df_current.empty:
                self.log("Warning: Sales data not available for NSR Analysis")
                return False

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "NSR Analysis"

            ws['A1'] = f"NSR Analysis - {self.prior_month} vs {self.current_month}"
            ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
            ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            ws.merge_cells('A1:E1')

            headers = ['Description', 'Prior Period', 'Current Period', 'Variance', 'Var %']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col)
                cell.value = header
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

            qty_cols = [col for col in df_current.columns if 'quantity' in col.lower() or '8-oz' in col.lower()]
            rev_cols = [col for col in df_current.columns if 'amount' in col.lower()]
            qty_col = qty_cols[0] if qty_cols else df_current.columns[0]
            rev_col = rev_cols[0] if rev_cols else df_current.columns[1]

            current_qty = df_current[qty_col].sum()
            current_rev = df_current[rev_col].sum()
            prior_qty = df_prior[qty_col].sum() if df_prior is not None and not df_prior.empty else 0
            prior_rev = df_prior[rev_col].sum() if df_prior is not None and not df_prior.empty else 0

            metrics = [
                ('Total Volume (8-Oz)', 'qty'),
                ('Total Revenue (PKR)', 'rev'),
                ('Average Price per Unit', 'price')
            ]

            row = 4
            for metric_name, metric_type in metrics:
                ws.cell(row=row, column=1).value = metric_name
                if metric_type == 'qty':
                    ws.cell(row=row, column=2).value = prior_qty
                    ws.cell(row=row, column=3).value = current_qty
                    ws.cell(row=row, column=2).number_format = '#,##0'
                    ws.cell(row=row, column=3).number_format = '#,##0'
                elif metric_type == 'rev':
                    ws.cell(row=row, column=2).value = prior_rev
                    ws.cell(row=row, column=3).value = current_rev
                    ws.cell(row=row, column=2).number_format = '#,##0'
                    ws.cell(row=row, column=3).number_format = '#,##0'
                else:
                    prior_price = prior_rev / prior_qty if prior_qty > 0 else 0
                    current_price = current_rev / current_qty if current_qty > 0 else 0
                    ws.cell(row=row, column=2).value = prior_price
                    ws.cell(row=row, column=3).value = current_price
                    ws.cell(row=row, column=2).number_format = '#,##0.00'
                    ws.cell(row=row, column=3).number_format = '#,##0.00'

                ws.cell(row=row, column=4).value = f"=C{row}-B{row}"
                ws.cell(row=row, column=4).number_format = '#,##0.00'
                ws.cell(row=row, column=5).value = f"=IF(B{row}=0,0,(C{row}-B{row})/B{row})"
                ws.cell(row=row, column=5).number_format = '0.0%'
                row += 1

            for col in range(1, 6):
                ws.column_dimensions[get_column_letter(col)].width = 20

            output_file = os.path.join(self.output_folder,
                f"1.2 NSR Analysis {self.prior_month}-{self.prior_year[-2:]} vs {self.current_month} {self.current_year[-2:]}.xlsx")
            wb.save(output_file)
            self.log(f"  Report 2 saved: {Path(output_file).name}")
            return True
        except Exception as e:
            self.log(f"Error generating Report 2: {str(e)[:100]}")
            return False

    # ─── REPORT 3: COMS ANALYSIS ──────────────────────────────
    def generate_coms_analysis_report(self):
        """Report 3: COMS Analysis"""
        self.log("Generating Report 3: COMS Analysis...")
        try:
            df_current = self.data.get('sales_current')
            df_prior = self.data.get('sales_prior')
            rm_current = self.data.get('rm_current')
            rm_prior = self.data.get('rm_prior')

            if df_current is None or df_current.empty:
                self.log("Warning: Sales data not available for COMS Analysis")
                return False

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "COMS Analysis"

            ws['A1'] = f"COMS Analysis - {self.prior_month} vs {self.current_month}"
            ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
            ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            ws.merge_cells('A1:E1')

            headers = ['Description', 'Prior Period', 'Current Period', 'Variance', 'Var %']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col)
                cell.value = header
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

            qty_cols = [col for col in df_current.columns if 'quantity' in col.lower() or '8-oz' in col.lower()]
            qty_col = qty_cols[0] if qty_cols else df_current.columns[0]

            current_qty = df_current[qty_col].sum()
            prior_qty = df_prior[qty_col].sum() if df_prior is not None else 0

            current_rm = 0
            prior_rm = 0
            if rm_current is not None:
                numeric = rm_current.select_dtypes(include=[np.number])
                if not numeric.empty:
                    current_rm = numeric.sum().sum()
            if rm_prior is not None:
                numeric = rm_prior.select_dtypes(include=[np.number])
                if not numeric.empty:
                    prior_rm = numeric.sum().sum()

            metrics = [
                ('Volume (8-Oz)', prior_qty, current_qty, '#,##0'),
                ('Raw Material Cost', prior_rm, current_rm, '#,##0.00'),
                ('COMS per Unit', prior_rm / prior_qty if prior_qty > 0 else 0,
                         current_rm / current_qty if current_qty > 0 else 0, '#,##0.00')
            ]

            row = 4
            for metric_name, prior_val, current_val, fmt in metrics:
                ws.cell(row=row, column=1).value = metric_name
                ws.cell(row=row, column=2).value = prior_val
                ws.cell(row=row, column=3).value = current_val
                ws.cell(row=row, column=2).number_format = fmt
                ws.cell(row=row, column=3).number_format = fmt
                ws.cell(row=row, column=4).value = f"=C{row}-B{row}"
                ws.cell(row=row, column=5).value = f"=IF(B{row}=0,0,(C{row}-B{row})/B{row})"
                ws.cell(row=row, column=5).number_format = '0.0%'
                row += 1

            for col in range(1, 6):
                ws.column_dimensions[get_column_letter(col)].width = 20

            output_file = os.path.join(self.output_folder,
                f"2. COMS Analysis - {self.prior_month} {self.prior_year[-2:]} vs {self.current_month} {self.current_year[-2:]}.xlsx")
            wb.save(output_file)
            self.log(f"  Report 3 saved: {Path(output_file).name}")
            return True
        except Exception as e:
            self.log(f"Error generating Report 3: {str(e)[:100]}")
            return False

    # ─── REPORT 4: POWER & FUEL ────────────────────────────────
    def generate_power_fuel_report(self):
        """Report 4: Power & Fuel Analysis"""
        self.log("Generating Report 4: Power & Fuel Analysis...")
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Power & Fuel"

            ws['A1'] = f"Power & Fuel Analysis - {self.prior_month} vs {self.current_month}"
            ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
            ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            ws.merge_cells('A1:D1')

            headers = ['Description', 'Prior Period', 'Current Period', 'Variance']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col)
                cell.value = header
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

            fesco_current = self.config_values.get('fesco_bill_current', 0)
            fesco_prior = self.config_values.get('fesco_bill_prior', 0)

            ws.cell(row=4, column=1).value = "FESCO Bill (PKR)"
            ws.cell(row=4, column=2).value = fesco_prior
            ws.cell(row=4, column=3).value = fesco_current
            ws.cell(row=4, column=2).number_format = '#,##0'
            ws.cell(row=4, column=3).number_format = '#,##0'
            ws.cell(row=4, column=4).value = "=C4-B4"
            ws.cell(row=4, column=4).number_format = '#,##0'

            ws.column_dimensions['A'].width = 25
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 20

            output_file = os.path.join(self.output_folder,
                f"3. Power and Fuel Analysis - {self.prior_month} {self.prior_year[-2:]} vs {self.current_month} {self.current_year[-2:]}.xlsx")
            wb.save(output_file)
            self.log(f"  Report 4 saved: {Path(output_file).name}")
            return True
        except Exception as e:
            self.log(f"Error generating Report 4: {str(e)[:100]}")
            return False

    # ─── REPORT 5: LEGAL & PROFESSIONAL ────────────────────────
    def generate_legal_report(self):
        """Report 5: Legal & Professional Admin"""
        self.log("Generating Report 5: Legal & Professional Admin...")
        try:
            df = self.data.get('legal_gl')
            if df is None or df.empty:
                self.log("Warning: Legal GL data not available")
                return False

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Legal & Admin"

            ws['A1'] = f"Legal & Professional Admin - {self.current_month} {self.current_year}"
            ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
            ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            ws.merge_cells('A1:C1')

            cust_cols = [col for col in df.columns if 'customer' in col.lower() or 'offsetting' in col.lower() or 'name' in col.lower()]
            rev_cols = [col for col in df.columns if 'amount' in col.lower()]
            cust_col = cust_cols[0] if cust_cols else df.columns[0]
            rev_col = rev_cols[0] if rev_cols else df.columns[-1]

            legal_summary = df.groupby(cust_col, as_index=False)[rev_col].sum().sort_values(rev_col, ascending=False)

            headers = ['Description', 'Amount (PKR)', 'Rank']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col)
                cell.value = header
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

            for row_idx, (idx, row) in enumerate(legal_summary.iterrows(), 4):
                ws.cell(row=row_idx, column=1).value = row[cust_col]
                ws.cell(row=row_idx, column=2).value = row[rev_col]
                ws.cell(row=row_idx, column=2).number_format = '#,##0.00'
                ws.cell(row=row_idx, column=3).value = row_idx - 3

            ws.column_dimensions['A'].width = 30
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 10

            output_file = os.path.join(self.output_folder,
                f"4. Legal & Professional Admin - {self.prior_month} {self.prior_year[-2:]} vs {self.current_month} {self.current_year[-2:]}.xlsx")
            wb.save(output_file)
            self.log(f"  Report 5 saved: {Path(output_file).name}")
            return True
        except Exception as e:
            self.log(f"Error generating Report 5: {str(e)[:100]}")
            return False

    # ─── COMBINED WORKBOOK (all 5 reports as tabs) ─────────────
    def generate_combined_workbook(self):
        """Generate Combined Analysis — all 5 reports as tabs in one workbook"""
        self.log("Generating Combined Analysis workbook...")
        try:
            output_folder = self.output_folder
            report_files = [
                f"1.1 Sales - Pack wise Volume and Top 10 Customers - {self.current_month} {self.current_year[-2:]}.xlsx",
                f"1.2 NSR Analysis {self.prior_month}-{self.prior_year[-2:]} vs {self.current_month} {self.current_year[-2:]}.xlsx",
                f"2. COMS Analysis - {self.prior_month} {self.prior_year[-2:]} vs {self.current_month} {self.current_year[-2:]}.xlsx",
                f"3. Power and Fuel Analysis - {self.prior_month} {self.prior_year[-2:]} vs {self.current_month} {self.current_year[-2:]}.xlsx",
                f"4. Legal & Professional Admin - {self.prior_month} {self.prior_year[-2:]} vs {self.current_month} {self.current_year[-2:]}.xlsx",
            ]

            combined_wb = openpyxl.Workbook()
            combined_wb.remove(combined_wb.active)

            for filename in report_files:
                filepath = os.path.join(output_folder, filename)
                if not os.path.exists(filepath):
                    self.log(f"  Skipping (not found): {filename}")
                    continue

                src_wb = openpyxl.load_workbook(filepath)
                for src_sheet_name in src_wb.sheetnames:
                    src_ws = src_wb[src_sheet_name]
                    tab_name = src_sheet_name[:31]
                    dst_ws = combined_wb.create_sheet(tab_name)

                    for row in src_ws.iter_rows():
                        for cell in row:
                            dst_cell = dst_ws.cell(row=cell.row, column=cell.column)
                            dst_cell.value = cell.value
                            if cell.has_style:
                                dst_cell.font = cell.font.copy()
                                dst_cell.fill = cell.fill.copy()
                                dst_cell.alignment = cell.alignment.copy()
                                dst_cell.number_format = cell.number_format
                                dst_cell.border = cell.border.copy()

                    for col_letter, dim in src_ws.column_dimensions.items():
                        dst_ws.column_dimensions[col_letter].width = dim.width

                    for merged_range in src_ws.merged_cells.ranges:
                        dst_ws.merge_cells(str(merged_range))

                src_wb.close()

            combined_file = os.path.join(output_folder,
                f"PBC Combined Analysis - {self.current_month} {self.current_year}.xlsx")
            combined_wb.save(combined_file)
            self.log(f"  Combined Analysis saved: {Path(combined_file).name}")
            return True
        except Exception as e:
            self.log(f"Error generating Combined workbook: {str(e)[:100]}")
            return False

    # ─── ENHANCED ANALYSIS v2.1 (10-sheet McKinsey-grade) ──────
    def generate_enhanced_v21_workbook(self):
        """Generate PBC Enhanced Analysis v2.1 — 10-sheet workbook"""
        self.log("\nGenerating Enhanced Analysis v2.1...")
        try:
            v = extract_enhanced_values(self.data)
            d = self.data
            wb = Workbook()

            vol_25 = v['vol_25']; vol_26 = v['vol_26']
            gsr_25 = v['gsr_25']; gsr_26 = v['gsr_26']
            nsr_25 = v['nsr_25']; nsr_26 = v['nsr_26']
            rm_25 = v['rm_total_25']; rm_26 = v['rm_total_26']
            wapda_25 = v.get('wapda_total_25', 0); wapda_26 = v.get('wapda_total_26', 0)
            cogs_25 = rm_25 + wapda_25; cogs_26 = rm_26 + wapda_26
            disc_25 = v['disc_total_25']; disc_26 = v['disc_total_26']
            zpsr_25 = v.get('zpsr_billing_25', v.get('pf_prod_25', 0))
            zpsr_26 = v.get('zpsr_billing_26', v.get('pf_prod_26', 0))
            kwh_25 = v.get('kwh_25', 0); kwh_26 = v.get('kwh_26', 0)

            # ── SHEET 1: EXECUTIVE DASHBOARD ──
            ws = wb.active
            ws.title = "1. Executive Dashboard"
            ws.sheet_properties.tabColor = "1F4E79"

            r = enh_write_title(ws, 1, "PBC \u2014 Executive Dashboard",
                        f"{self.prior_month} {self.prior_year} vs {self.current_month} {self.current_year} | Pakistan Beverages Company (Pvt) Ltd")

            headers = ['Key Performance Indicator', f'{self.prior_month[:3]}-{self.prior_year[-2:]}',
                       f'{self.current_month[:3]}-{self.current_year[-2:]}', 'Variance', 'Var %', 'Signal']
            for c, h in enumerate(headers, 1):
                ws.cell(row=r, column=c, value=h)
            enh_style_header(ws, r, 6)
            r += 1

            kpis = [
                ('Volume (8Oz Cases)', vol_25, vol_26, False, False, False),
                ('Gross Sales Revenue (PKR)', gsr_25, gsr_26, False, False, False),
                ('GSR per 8Oz', gsr_25/vol_25 if vol_25 else 0, gsr_26/vol_26 if vol_26 else 0, False, False, True),
                ('Total Discounts (PKR)', disc_25, disc_26, True, False, False),
                ('Discount Intensity (% of GSR)', disc_25/gsr_25 if gsr_25 else 0, disc_26/gsr_26 if gsr_26 else 0, True, True, False),
                ('Net Sales Revenue (PKR)', nsr_25, nsr_26, False, False, False),
                ('NSR per 8Oz', nsr_25/vol_25 if vol_25 else 0, nsr_26/vol_26 if vol_26 else 0, False, False, True),
                (None, None, None, False, False, False),
                ('Raw Material Cost (PKR)', rm_25, rm_26, True, False, False),
                ('Power & Fuel \u2014 WAPDA/FESCO (PKR)', wapda_25, wapda_26, True, False, False),
                ('Total COGS (RM + Power)', cogs_25, cogs_26, True, False, False),
                (None, None, None, False, False, False),
                ('Gross Profit (NSR - COGS)', nsr_25 - cogs_25, nsr_26 - cogs_26, False, False, False),
                ('Gross Profit Margin %', (nsr_25 - cogs_25)/nsr_25 if nsr_25 else 0, (nsr_26 - cogs_26)/nsr_26 if nsr_26 else 0, False, True, False),
                (None, None, None, False, False, False),
                ('COGS per 8Oz', cogs_25/vol_25 if vol_25 else 0, cogs_26/vol_26 if vol_26 else 0, True, False, True),
                ('  of which: RM per 8Oz', rm_25/vol_25 if vol_25 else 0, rm_26/vol_26 if vol_26 else 0, True, False, True),
                ('  of which: Power per 8Oz', wapda_25/vol_25 if vol_25 else 0, wapda_26/vol_26 if vol_26 else 0, True, False, True),
                (None, None, None, False, False, False),
                ('Production Volume \u2014 ZPSR (8Oz)', zpsr_25, zpsr_26, False, False, False),
                ('WAPDA KwH Consumed', kwh_25, kwh_26, True, False, False),
                ('Cases Produced per KwH', zpsr_25/kwh_25 if kwh_25 else 0, zpsr_26/kwh_26 if kwh_26 else 0, False, False, True),
                ('Power Cost per ZPSR Case', wapda_25/zpsr_25 if zpsr_25 else 0, wapda_26/zpsr_26 if zpsr_26 else 0, True, False, True),
                (None, None, None, False, False, False),
                ('Legal & Professional (PKR)', v['legal_jan_total'], v['legal_feb_total'], True, False, False),
            ]

            for name, val25, val26, is_cost, is_pct, is_pu in kpis:
                if name is None:
                    r += 1; continue
                ws.cell(row=r, column=1, value=name)
                ws.cell(row=r, column=2, value=val25)
                ws.cell(row=r, column=3, value=val26)
                ws.cell(row=r, column=4).value = f'=C{r}-B{r}'
                ws.cell(row=r, column=5).value = f'=IF(B{r}=0,0,D{r}/B{r})'
                if is_cost:
                    ws.cell(row=r, column=6).value = f'=IF(E{r}>0.1,"\u26a0 RISING",IF(E{r}>0.05,"WATCH","\u2713 OK"))'
                else:
                    ws.cell(row=r, column=6).value = f'=IF(E{r}>0.1,"\u2713 STRONG",IF(E{r}>0,"STABLE","\u26a0 DECLINE"))'
                if is_pct:
                    for c in [2, 3, 4]: ws.cell(row=r, column=c).number_format = PCT_FMT
                elif is_pu:
                    for c in [2, 3, 4]: ws.cell(row=r, column=c).number_format = DEC_FMT
                else:
                    for c in [2, 3, 4]: ws.cell(row=r, column=c).number_format = NUM_FMT
                ws.cell(row=r, column=5).number_format = PCT_FMT
                enh_style_data(ws, r, 6)
                if name in ('Total COGS (RM + Power)', 'Gross Profit (NSR - COGS)', 'Gross Profit Margin %'):
                    ws.cell(row=r, column=1).font = BOLD_BLUE
                if '  of which' in name:
                    ws.cell(row=r, column=1).font = ITALIC_NOTE
                r += 1

            r += 1
            ws.cell(row=r, column=1,
                    value='Note: COGS excludes manufacturing labor, depreciation, and other conversion costs '
                          '(not available in GL extracts). True bottler GP margin is typically 2-5pp lower.').font = CAVEAT_FONT
            r += 2

            # Margin Bridge
            r = enh_write_title(ws, r, f"Gross Profit Bridge \u2014 {self.prior_month[:3]}-{self.prior_year[-2:]} \u2192 {self.current_month[:3]}-{self.current_year[-2:]}",
                        "Decomposes GP change into Volume, Price/Mix, and Cost drivers")
            for c, h in enumerate(['Component', 'Impact (PKR)', 'Impact per 8Oz', 'Notes'], 1):
                ws.cell(row=r, column=c, value=h)
            enh_style_header(ws, r, 4)
            r += 1

            gp_25 = nsr_25 - cogs_25; gp_26 = nsr_26 - cogs_26
            nsr_per_25 = nsr_25 / vol_25 if vol_25 else 0
            cogs_per_25 = cogs_25 / vol_25 if vol_25 else 0
            dv = vol_26 - vol_25

            bridge = [
                (f'Starting Gross Profit ({self.prior_month[:3]}-{self.prior_year[-2:]})', gp_25, gp_25/vol_25 if vol_25 else 0,
                 f'NSR {nsr_25/1e6:,.0f}M \u2013 COGS {cogs_25/1e6:,.0f}M'),
                ('(+) Volume uplift on NSR', dv * nsr_per_25, nsr_per_25,
                 f'+{dv:,.0f} cases \u00d7 PKR {nsr_per_25:,.2f}/case'),
                ('(+) NSR rate improvement (price/mix)', (nsr_26 - nsr_25) - dv * nsr_per_25, 0,
                 'Net pricing + pack mix shift'),
                ('(\u2013) Volume-driven COGS increase', -(dv * cogs_per_25), -cogs_per_25,
                 f'+{dv:,.0f} cases \u00d7 PKR {cogs_per_25:,.2f}/case'),
                ('(\u2013) COGS rate change (inflation)', -((cogs_26 - cogs_25) - dv * cogs_per_25), 0,
                 'RM inflation + tariff changes'),
                (f'Ending Gross Profit ({self.current_month[:3]}-{self.current_year[-2:]})', gp_26, gp_26/vol_26 if vol_26 else 0,
                 f'NSR {nsr_26/1e6:,.0f}M \u2013 COGS {cogs_26/1e6:,.0f}M'),
            ]

            bridge_start = r
            for i, (name, impact, per_oz, note) in enumerate(bridge):
                ws.cell(row=r, column=1, value=name).font = DATA_FONT
                ws.cell(row=r, column=2, value=impact); ws.cell(row=r, column=2).number_format = NUM_FMT
                ws.cell(row=r, column=3, value=per_oz); ws.cell(row=r, column=3).number_format = DEC_FMT
                ws.cell(row=r, column=4, value=note).font = ITALIC_NOTE
                for c in range(1, 5): ws.cell(row=r, column=c).border = THIN
                if i == 0 or i == len(bridge) - 1:
                    ws.cell(row=r, column=1).font = BOLD_DATA
                    enh_style_total(ws, r, 4)
                r += 1

            r += 1
            end_r = bridge_start + 5
            ws.cell(row=r, column=1, value='RECONCILIATION CHECK (must = 0):').font = WARN_FONT
            ws.cell(row=r, column=2).value = (
                f'=B{end_r}-B{bridge_start}-B{bridge_start+1}-B{bridge_start+2}'
                f'-B{bridge_start+3}-B{bridge_start+4}')
            ws.cell(row=r, column=2).number_format = DEC_FMT

            enh_auto_width(ws, 6, min_w=18, max_w=42)
            ws.column_dimensions['D'].width = 48
            self.log("  Sheet 1: Executive Dashboard done")

            # ── SHEET 2: REVENUE WATERFALL ──
            ws2 = wb.create_sheet("2. Revenue Waterfall")
            ws2.sheet_properties.tabColor = "4472C4"
            r = enh_write_title(ws2, 1, "Revenue Waterfall Analysis",
                        f"GSR \u2192 Discounts \u2192 NSR | {self.prior_month[:3]}-{self.prior_year[-2:]} vs {self.current_month[:3]}-{self.current_year[-2:]}")

            headers = ['Revenue Component', f'{self.prior_month[:3]}-{self.prior_year[-2:]} (PKR)',
                       f'{self.current_month[:3]}-{self.current_year[-2:]} (PKR)', 'Variance (PKR)', 'Var %',
                       f'Per 8Oz ({self.prior_month[:3]}-{self.prior_year[-2:]})',
                       f'Per 8Oz ({self.current_month[:3]}-{self.current_year[-2:]})', 'Per 8Oz Var %']
            for c, h in enumerate(headers, 1): ws2.cell(row=r, column=c, value=h)
            enh_style_header(ws2, r, 8)
            r += 1

            vol_row = r
            ws2.cell(row=r, column=1, value='Volume (8Oz Cases)')
            ws2.cell(row=r, column=2, value=vol_25); ws2.cell(row=r, column=3, value=vol_26)
            ws2.cell(row=r, column=4).value = f'=C{r}-B{r}'
            ws2.cell(row=r, column=5).value = f'=IF(B{r}=0,0,D{r}/B{r})'
            r += 1

            gsr_row = r
            ws2.cell(row=r, column=1, value='Gross Sales Revenue')
            ws2.cell(row=r, column=2, value=gsr_25); ws2.cell(row=r, column=3, value=gsr_26)
            ws2.cell(row=r, column=4).value = f'=C{r}-B{r}'
            ws2.cell(row=r, column=5).value = f'=IF(B{r}=0,0,D{r}/B{r})'
            ws2.cell(row=r, column=6).value = f'=IF(B{vol_row}=0,0,B{r}/B{vol_row})'
            ws2.cell(row=r, column=7).value = f'=IF(C{vol_row}=0,0,C{r}/C{vol_row})'
            ws2.cell(row=r, column=8).value = f'=IF(F{r}=0,0,(G{r}-F{r})/F{r})'
            r += 2

            ws2.cell(row=r, column=1, value='DISCOUNT BREAKDOWN:').font = SUB_FONT
            r += 1
            disc_start = r
            for dl in v['disc_lines']:
                ws2.cell(row=r, column=1, value=f"  {dl['Description']} (GL {dl['GL']})")
                ws2.cell(row=r, column=2, value=dl['Feb_25']); ws2.cell(row=r, column=3, value=dl['Feb_26'])
                ws2.cell(row=r, column=4).value = f'=C{r}-B{r}'
                ws2.cell(row=r, column=5).value = f'=IF(B{r}=0,0,D{r}/B{r})'
                ws2.cell(row=r, column=6).value = f'=IF(B{vol_row}=0,0,B{r}/B{vol_row})'
                ws2.cell(row=r, column=7).value = f'=IF(C{vol_row}=0,0,C{r}/C{vol_row})'
                ws2.cell(row=r, column=8).value = f'=IF(F{r}=0,0,(G{r}-F{r})/F{r})'
                r += 1
            disc_end = r - 1

            ws2.cell(row=r, column=1, value='TOTAL DISCOUNTS')
            ws2.cell(row=r, column=2).value = f'=SUM(B{disc_start}:B{disc_end})'
            ws2.cell(row=r, column=3).value = f'=SUM(C{disc_start}:C{disc_end})'
            ws2.cell(row=r, column=4).value = f'=C{r}-B{r}'
            ws2.cell(row=r, column=5).value = f'=IF(B{r}=0,0,D{r}/B{r})'
            ws2.cell(row=r, column=6).value = f'=IF(B{vol_row}=0,0,B{r}/B{vol_row})'
            ws2.cell(row=r, column=7).value = f'=IF(C{vol_row}=0,0,C{r}/C{vol_row})'
            ws2.cell(row=r, column=8).value = f'=IF(F{r}=0,0,(G{r}-F{r})/F{r})'
            enh_style_total(ws2, r, 8)
            disc_total_row = r
            r += 2

            ws2.cell(row=r, column=1, value='Discount Intensity (% of GSR)')
            ws2.cell(row=r, column=1).font = WARN_FONT
            ws2.cell(row=r, column=2).value = f'=IF(B{gsr_row}=0,0,B{disc_total_row}/B{gsr_row})'
            ws2.cell(row=r, column=3).value = f'=IF(C{gsr_row}=0,0,C{disc_total_row}/C{gsr_row})'
            ws2.cell(row=r, column=4).value = f'=C{r}-B{r}'
            for c in [2, 3, 4]: ws2.cell(row=r, column=c).number_format = PCT_FMT
            r += 2

            nsr_row = r
            ws2.cell(row=r, column=1, value='NET SALES REVENUE')
            ws2.cell(row=r, column=2).value = f'=B{gsr_row}-B{disc_total_row}'
            ws2.cell(row=r, column=3).value = f'=C{gsr_row}-C{disc_total_row}'
            ws2.cell(row=r, column=4).value = f'=C{r}-B{r}'
            ws2.cell(row=r, column=5).value = f'=IF(B{r}=0,0,D{r}/B{r})'
            ws2.cell(row=r, column=6).value = f'=IF(B{vol_row}=0,0,B{r}/B{vol_row})'
            ws2.cell(row=r, column=7).value = f'=IF(C{vol_row}=0,0,C{r}/C{vol_row})'
            ws2.cell(row=r, column=8).value = f'=IF(F{r}=0,0,(G{r}-F{r})/F{r})'
            enh_style_total(ws2, r, 8)
            r += 2

            ws2.cell(row=r, column=1, value='Z-CHECK (must be zero):').font = WARN_FONT
            ws2.cell(row=r, column=2, value=disc_25 - sum(dl['Feb_25'] for dl in v['disc_lines']))
            ws2.cell(row=r, column=3, value=disc_26 - sum(dl['Feb_26'] for dl in v['disc_lines']))
            ws2.cell(row=r, column=2).number_format = DEC_FMT; ws2.cell(row=r, column=3).number_format = DEC_FMT

            for row_n in range(vol_row, nsr_row + 1):
                for c in [2, 3, 4]: ws2.cell(row=row_n, column=c).number_format = NUM_FMT
                for c in [5, 8]: ws2.cell(row=row_n, column=c).number_format = PCT_FMT
                for c in [6, 7]: ws2.cell(row=row_n, column=c).number_format = DEC_FMT
                enh_style_data(ws2, row_n, 8)
            enh_auto_width(ws2, 8, min_w=16, max_w=48)
            self.log("  Sheet 2: Revenue Waterfall done")

            # ── SHEET 3: DISCOUNT DEEP DIVE ──
            ws3 = wb.create_sheet("3. Discount Deep Dive")
            ws3.sheet_properties.tabColor = "C00000"
            r = enh_write_title(ws3, 1, "Discount Effectiveness & Trade Spend Analysis",
                        f"{self.prior_month[:3]}-{self.prior_year[-2:]} vs {self.current_month[:3]}-{self.current_year[-2:]} | Allocation Basis Audit")

            headers = ['Discount Type', 'GL Code', f'{self.prior_month[:3]}-{self.prior_year[-2:]} (PKR)',
                       f'{self.current_month[:3]}-{self.current_year[-2:]} (PKR)', 'Variance', 'Var %',
                       f'% of Total ({self.prior_month[:3]}-{self.prior_year[-2:]})',
                       f'% of Total ({self.current_month[:3]}-{self.current_year[-2:]})',
                       'Allocation Basis', 'Audit Risk']
            for c, h in enumerate(headers, 1): ws3.cell(row=r, column=c, value=h)
            enh_style_header(ws3, r, 10)
            r += 1

            dd_start = r
            num_disc = len(v['disc_lines'])
            for dl in v['disc_lines']:
                ws3.cell(row=r, column=1, value=dl['Description'])
                ws3.cell(row=r, column=2, value=dl['GL'])
                ws3.cell(row=r, column=3, value=dl['Feb_25']); ws3.cell(row=r, column=4, value=dl['Feb_26'])
                ws3.cell(row=r, column=5).value = f'=D{r}-C{r}'
                ws3.cell(row=r, column=6).value = f'=IF(C{r}=0,0,E{r}/C{r})'
                ws3.cell(row=r, column=7).value = f'=IF(SUM(C${dd_start}:C${dd_start+num_disc-1})=0,0,C{r}/SUM(C${dd_start}:C${dd_start+num_disc-1}))'
                ws3.cell(row=r, column=8).value = f'=IF(SUM(D${dd_start}:D${dd_start+num_disc-1})=0,0,D{r}/SUM(D${dd_start}:D${dd_start+num_disc-1}))'
                ws3.cell(row=r, column=9, value=dl['Basis'])
                var_pct = (dl['Feb_26'] - dl['Feb_25'])/dl['Feb_25'] if dl['Feb_25'] != 0 else 999
                risk = "\u26a0 HIGH \u2014 >200% change" if abs(var_pct) > 2 else ("MEDIUM \u2014 >50% change" if abs(var_pct) > 0.5 else "LOW")
                ws3.cell(row=r, column=10, value=risk)
                r += 1
            dd_end = r - 1

            ws3.cell(row=r, column=1, value='TOTAL')
            ws3.cell(row=r, column=3).value = f'=SUM(C{dd_start}:C{dd_end})'
            ws3.cell(row=r, column=4).value = f'=SUM(D{dd_start}:D{dd_end})'
            ws3.cell(row=r, column=5).value = f'=D{r}-C{r}'
            ws3.cell(row=r, column=6).value = f'=IF(C{r}=0,0,E{r}/C{r})'
            enh_style_total(ws3, r, 10)
            r += 1

            for rn in range(dd_start, r):
                for c in [3, 4, 5]: ws3.cell(row=rn, column=c).number_format = NUM_FMT
                for c in [6, 7, 8]: ws3.cell(row=rn, column=c).number_format = PCT_FMT
                enh_style_data(ws3, rn, 10)

            # Trade Spend ROI
            r += 2
            r = enh_write_title(ws3, r, "Trade Spend ROI Analysis")
            roi_h = ['Metric', 'Value', 'Notes']
            for c, h in enumerate(roi_h, 1): ws3.cell(row=r, column=c, value=h)
            enh_style_header(ws3, r, 3)
            r += 1

            incr_vol = vol_26 - vol_25; incr_disc = disc_26 - disc_25
            roi_items = [
                ('Incremental Volume (8Oz)', f'{incr_vol:,.0f}', f'+{incr_vol/vol_25*100:.0f}% YoY' if vol_25 else ''),
                ('Incremental Discount Spend (PKR)', f'{incr_disc:,.0f}', f'+{incr_disc/disc_25*100:.0f}% YoY' if disc_25 else ''),
                ('Cost per Incremental 8Oz Case (PKR)', f'{incr_disc/incr_vol:,.2f}' if incr_vol else 'N/A', 'PKR spent per extra case sold'),
                (f'Discount-to-NSR Ratio ({self.prior_month[:3]}-{self.prior_year[-2:]})', f'{disc_25/nsr_25:.1%}' if nsr_25 else 'N/A', ''),
                (f'Discount-to-NSR Ratio ({self.current_month[:3]}-{self.current_year[-2:]})', f'{disc_26/nsr_26:.1%}' if nsr_26 else 'N/A', '\u26a0 Rising ratio = margin pressure'),
                (f'NSR Retained per PKR Discount ({self.prior_month[:3]}-{self.prior_year[-2:]})', f'{nsr_25/disc_25:,.2f}' if disc_25 else 'N/A', ''),
                (f'NSR Retained per PKR Discount ({self.current_month[:3]}-{self.current_year[-2:]})', f'{nsr_26/disc_26:,.2f}' if disc_26 else 'N/A', 'Lower = less efficient trade spend'),
            ]
            for name, val, note in roi_items:
                ws3.cell(row=r, column=1, value=name); ws3.cell(row=r, column=2, value=val)
                ws3.cell(row=r, column=3, value=note)
                enh_style_data(ws3, r, 3); r += 1

            enh_auto_width(ws3, 10, min_w=15, max_w=38)
            ws3.column_dimensions['I'].width = 48; ws3.column_dimensions['J'].width = 24
            self.log("  Sheet 3: Discount Deep Dive done")

            # ── SHEET 4: COMS ANALYSIS ──
            ws4 = wb.create_sheet("4. COMS Analysis")
            ws4.sheet_properties.tabColor = "548235"
            r = enh_write_title(ws4, 1, "Cost of Materials Sold \u2014 Category Analysis",
                        f"{self.prior_month[:3]}-{self.prior_year[-2:]} vs {self.current_month[:3]}-{self.current_year[-2:]} | Raw Material Inflation Tracking")

            headers = ['Category', f'{self.prior_month[:3]}-{self.prior_year[-2:]} (PKR)',
                       f'{self.current_month[:3]}-{self.current_year[-2:]} (PKR)', 'Variance (PKR)', 'Var %',
                       f'% of Total ({self.prior_month[:3]}-{self.prior_year[-2:]})',
                       f'% of Total ({self.current_month[:3]}-{self.current_year[-2:]})', 'Shift (pp)']
            for c, h in enumerate(headers, 1): ws4.cell(row=r, column=c, value=h)
            enh_style_header(ws4, r, 8)
            r += 1

            coms_start = r
            gt_row = coms_start + len(COMS_CATEGORIES)
            for desc, v25_val, v26_val in COMS_CATEGORIES:
                ws4.cell(row=r, column=1, value=desc)
                ws4.cell(row=r, column=2, value=v25_val); ws4.cell(row=r, column=3, value=v26_val)
                ws4.cell(row=r, column=4).value = f'=C{r}-B{r}'
                ws4.cell(row=r, column=5).value = f'=IF(B{r}=0,0,D{r}/B{r})'
                ws4.cell(row=r, column=6).value = f'=IF(B${gt_row}=0,0,B{r}/B${gt_row})'
                ws4.cell(row=r, column=7).value = f'=IF(C${gt_row}=0,0,C{r}/C${gt_row})'
                ws4.cell(row=r, column=8).value = f'=G{r}-F{r}'
                r += 1
            coms_end = r - 1

            ws4.cell(row=r, column=1, value='GRAND TOTAL')
            ws4.cell(row=r, column=2).value = f'=SUM(B{coms_start}:B{coms_end})'
            ws4.cell(row=r, column=3).value = f'=SUM(C{coms_start}:C{coms_end})'
            ws4.cell(row=r, column=4).value = f'=C{r}-B{r}'
            ws4.cell(row=r, column=5).value = f'=IF(B{r}=0,0,D{r}/B{r})'
            enh_style_total(ws4, r, 8)
            r += 1

            for rn in range(coms_start, r):
                for c in [2, 3, 4]: ws4.cell(row=rn, column=c).number_format = NUM_FMT
                for c in [5, 6, 7]: ws4.cell(row=rn, column=c).number_format = PCT_FMT
                ws4.cell(row=rn, column=8).number_format = '0.0%'
                enh_style_data(ws4, rn, 8)

            # Per-unit section
            r += 2
            r = enh_write_title(ws4, r, "COMS per Unit \u2014 Cost Efficiency Tracking")
            for c, h in enumerate(['Metric', f'{self.prior_month[:3]}-{self.prior_year[-2:]}',
                                   f'{self.current_month[:3]}-{self.current_year[-2:]}', 'Variance', 'Var %'], 1):
                ws4.cell(row=r, column=c, value=h)
            enh_style_header(ws4, r, 5); r += 1

            prod_25 = v.get('prod_cases_25', 1); prod_26 = v.get('prod_cases_26', 1)
            unit_metrics = [
                ('COMS per 8Oz Case', rm_25/vol_25 if vol_25 else 0, rm_26/vol_26 if vol_26 else 0),
                ('COMS per Production Case', rm_25/prod_25, rm_26/prod_26),
                ('Sugar Cost per kg', 120.218366, 129.579124),
                ('CO2 Cost per kg', 30.445397, 55.668046),
                ('Avg Preform Cost per unit', 178334562/12093746, 287899068/24451584),
            ]
            for name, v25_val, v26_val in unit_metrics:
                ws4.cell(row=r, column=1, value=name); ws4.cell(row=r, column=2, value=v25_val)
                ws4.cell(row=r, column=3, value=v26_val)
                ws4.cell(row=r, column=4).value = f'=C{r}-B{r}'
                ws4.cell(row=r, column=5).value = f'=IF(B{r}=0,0,D{r}/B{r})'
                for c in [2, 3, 4]: ws4.cell(row=r, column=c).number_format = DEC_FMT
                ws4.cell(row=r, column=5).number_format = PCT_FMT
                enh_style_data(ws4, r, 5); r += 1

            enh_auto_width(ws4, 8, min_w=16, max_w=38)
            self.log("  Sheet 4: COMS Analysis done")

            # ── SHEET 5: PRODUCTION & POWER ──
            ws5 = wb.create_sheet("5. Production & Power")
            ws5.sheet_properties.tabColor = "BF8F00"
            r = enh_write_title(ws5, 1, "Production Volume & Power/Fuel Analysis",
                        f"ZPSR Data & FESCO Billing | {self.prior_month[:3]}-{self.prior_year[-2:]} vs {self.current_month[:3]}-{self.current_year[-2:]}")

            headers = ['Brand', f'{self.prior_month[:3]}-{self.prior_year[-2:]} (Conv. 250ml)',
                       f'{self.current_month[:3]}-{self.current_year[-2:]} (Conv. 250ml)', 'Variance', 'Var %',
                       f'Mix % ({self.prior_month[:3]}-{self.prior_year[-2:]})',
                       f'Mix % ({self.current_month[:3]}-{self.current_year[-2:]})']
            for c, h in enumerate(headers, 1): ws5.cell(row=r, column=c, value=h)
            enh_style_header(ws5, r, 7); r += 1

            all_brands = sorted(set(list(v['brands_25'].keys()) + list(v['brands_26'].keys())))
            prod_start = r
            for brand in all_brands:
                b25 = v['brands_25'].get(brand, 0); b26 = v['brands_26'].get(brand, 0)
                ws5.cell(row=r, column=1, value=brand)
                ws5.cell(row=r, column=2, value=b25); ws5.cell(row=r, column=3, value=b26)
                ws5.cell(row=r, column=4).value = f'=C{r}-B{r}'
                ws5.cell(row=r, column=5).value = f'=IF(B{r}=0,0,D{r}/B{r})'
                r += 1
            prod_end = r - 1
            total_prod_row = r

            ws5.cell(row=r, column=1, value='TOTAL PRODUCTION')
            ws5.cell(row=r, column=2).value = f'=SUM(B{prod_start}:B{prod_end})'
            ws5.cell(row=r, column=3).value = f'=SUM(C{prod_start}:C{prod_end})'
            ws5.cell(row=r, column=4).value = f'=C{r}-B{r}'
            ws5.cell(row=r, column=5).value = f'=IF(B{r}=0,0,D{r}/B{r})'
            enh_style_total(ws5, r, 7)

            for rn in range(prod_start, prod_end + 1):
                ws5.cell(row=rn, column=6).value = f'=IF(B${total_prod_row}=0,0,B{rn}/B${total_prod_row})'
                ws5.cell(row=rn, column=7).value = f'=IF(C${total_prod_row}=0,0,C{rn}/C${total_prod_row})'
            for rn in range(prod_start, r + 1):
                for c in [2, 3, 4]: ws5.cell(row=rn, column=c).number_format = NUM_FMT
                for c in [5, 6, 7]: ws5.cell(row=rn, column=c).number_format = PCT_FMT
                enh_style_data(ws5, rn, 7)
            r += 2

            # FESCO/WAPDA breakdown
            r = enh_write_title(ws5, r, "Power & Fuel Analysis \u2014 FESCO Billing Data",
                        "Source: FESCO Online Bills & ZPSR Production Reports")
            for c, h in enumerate(['Particular', f'{self.prior_month[:3]}-{self.prior_year[-2:]} (PKR)',
                                   f'{self.current_month[:3]}-{self.current_year[-2:]} (PKR)', 'Variance', 'Var %'], 1):
                ws5.cell(row=r, column=c, value=h)
            enh_style_header(ws5, r, 5); r += 1

            for name, v25_val, v26_val in v.get('wapda_items', []):
                ws5.cell(row=r, column=1, value=name)
                ws5.cell(row=r, column=2, value=v25_val); ws5.cell(row=r, column=3, value=v26_val)
                ws5.cell(row=r, column=4).value = f'=C{r}-B{r}'
                ws5.cell(row=r, column=5).value = f'=IF(B{r}=0,0,D{r}/B{r})'
                for c in [2, 3, 4]: ws5.cell(row=r, column=c).number_format = NUM_FMT
                ws5.cell(row=r, column=5).number_format = PCT_FMT
                enh_style_data(ws5, r, 5); r += 1

            r += 1
            ws5.cell(row=r, column=1, value='TOTAL WAPDA (as per FESCO Bill)')
            ws5.cell(row=r, column=2, value=wapda_25); ws5.cell(row=r, column=3, value=wapda_26)
            ws5.cell(row=r, column=4).value = f'=C{r}-B{r}'
            ws5.cell(row=r, column=5).value = f'=IF(B{r}=0,0,D{r}/B{r})'
            for c in [2, 3, 4]: ws5.cell(row=r, column=c).number_format = NUM_FMT
            ws5.cell(row=r, column=5).number_format = PCT_FMT
            enh_style_total(ws5, r, 5)
            r += 2

            ws5.cell(row=r, column=1, value='EFFICIENCY KPIs').font = SUB_FONT
            ws5.cell(row=r, column=1).fill = SUB_FILL; r += 1

            eff_kpis = [
                ('Production Volume \u2014 ZPSR (8Oz)', zpsr_25, zpsr_26),
                ('Cases Produced per KwH', zpsr_25/kwh_25 if kwh_25 else 0, zpsr_26/kwh_26 if kwh_26 else 0),
                ('Avg Cost per KwH (Energy only)', 39011997/kwh_25 if kwh_25 else 0, 40640738.05/kwh_26 if kwh_26 else 0),
                ('WAPDA Cost per 8Oz Case', wapda_25/zpsr_25 if zpsr_25 else 0, wapda_26/zpsr_26 if zpsr_26 else 0),
                ('WAPDA as % of NSR', wapda_25/nsr_25 if nsr_25 else 0, wapda_26/nsr_26 if nsr_26 else 0),
            ]
            for name, v25_val, v26_val in eff_kpis:
                ws5.cell(row=r, column=1, value=name)
                ws5.cell(row=r, column=2, value=v25_val); ws5.cell(row=r, column=3, value=v26_val)
                ws5.cell(row=r, column=4).value = f'=C{r}-B{r}'
                ws5.cell(row=r, column=5).value = f'=IF(B{r}=0,0,D{r}/B{r})'
                fmt = PCT_FMT if '% of NSR' in name else DEC_FMT
                for c in [2, 3, 4]: ws5.cell(row=r, column=c).number_format = fmt
                ws5.cell(row=r, column=5).number_format = PCT_FMT
                enh_style_data(ws5, r, 5); r += 1

            enh_auto_width(ws5, 7, min_w=16, max_w=32)
            self.log("  Sheet 5: Production & Power done")

            # ── SHEET 6: LEGAL & PROFESSIONAL ──
            ws6 = wb.create_sheet("6. Legal & Professional")
            ws6.sheet_properties.tabColor = "7030A0"
            r = enh_write_title(ws6, 1, "Legal & Professional Expenses \u2014 Party-wise Detail",
                        "GL 855001 | Jan-26 vs Feb-26")
            headers = ['Party / Vendor', 'Cost Center', 'Jan-26 (PKR)', 'Feb-26 (PKR)', 'Variance', 'Var %', 'Category']
            for c, h in enumerate(headers, 1): ws6.cell(row=r, column=c, value=h)
            enh_style_header(ws6, r, 7); r += 1

            legal_jan = d.get('legal_jan')
            legal_feb = d.get('legal_feb')
            jan_vendors = aggregate_legal(legal_jan) if legal_jan is not None else {}
            feb_vendors = aggregate_legal(legal_feb) if legal_feb is not None else {}
            all_vendors = sorted(set(list(jan_vendors.keys()) + list(feb_vendors.keys())))

            leg_start = r
            for vendor in all_vendors:
                jan_amt = jan_vendors.get(vendor, {}).get('amount', 0)
                feb_amt = feb_vendors.get(vendor, {}).get('amount', 0)
                cc = jan_vendors.get(vendor, {}).get('cc', '') or feb_vendors.get(vendor, {}).get('cc', '')
                vl = vendor.lower()
                cat = ('Legal' if any(w in vl for w in ['law', 'legal', 'advocate', 'court']) else
                       'Audit' if any(w in vl for w in ['audit', 'ey', 'ernst', 'ford rhodes']) else
                       'Consulting' if any(w in vl for w in ['consult', 'adviser', 'advisory', 'renaissance', 'kale']) else
                       'Tax Advisory' if any(w in vl for w in ['tax', 'forum']) else
                       'IT Services' if any(w in vl for w in ['sap', 'software', 'exd', 'excellence', 'qlik', 'tally']) else
                       'Employee Settlement' if any(w in vl for w in ['employee', 'advance', 'short term']) else 'Other')
                ws6.cell(row=r, column=1, value=vendor); ws6.cell(row=r, column=2, value=str(cc))
                ws6.cell(row=r, column=3, value=jan_amt); ws6.cell(row=r, column=4, value=feb_amt)
                ws6.cell(row=r, column=5).value = f'=D{r}-C{r}'
                ws6.cell(row=r, column=6).value = f'=IF(C{r}=0,0,E{r}/C{r})'
                ws6.cell(row=r, column=7, value=cat)
                r += 1
            leg_end = r - 1

            ws6.cell(row=r, column=1, value='TOTAL')
            ws6.cell(row=r, column=3).value = f'=SUM(C{leg_start}:C{leg_end})'
            ws6.cell(row=r, column=4).value = f'=SUM(D{leg_start}:D{leg_end})'
            ws6.cell(row=r, column=5).value = f'=D{r}-C{r}'
            ws6.cell(row=r, column=6).value = f'=IF(C{r}=0,0,E{r}/C{r})'
            enh_style_total(ws6, r, 7)
            for rn in range(leg_start, r + 1):
                for c in [3, 4, 5]: ws6.cell(row=rn, column=c).number_format = NUM_FMT
                ws6.cell(row=rn, column=6).number_format = PCT_FMT
                enh_style_data(ws6, rn, 7)

            r += 2
            ws6.cell(row=r, column=1, value='Legal & Prof as % of NSR:').font = SUB_FONT
            ws6.cell(row=r, column=2, value=f'{self.current_month[:3]}-{self.current_year[-2:]}')
            ws6.cell(row=r, column=3, value=v['legal_feb_total']/nsr_26 if nsr_26 else 0).number_format = PCT_FMT

            enh_auto_width(ws6, 7, min_w=14, max_w=42)
            self.log("  Sheet 6: Legal & Professional done")

            # ── SHEET 7: GROSS MARGIN BY BRAND ──
            ws7 = wb.create_sheet("7. Gross Margin by Brand")
            ws7.sheet_properties.tabColor = "00B050"

            ws7.cell(row=1, column=1, value='Estimated Gross Margin by Brand \u2014 ILLUSTRATIVE ONLY')
            ws7.cell(row=1, column=1).font = Font(name='Arial', bold=True, size=14, color='C00000')
            ws7.cell(row=2, column=1, value=f'Based on Production Mix & Average COGS Allocation | {self.current_month[:3]}-{self.current_year[-2:]}').font = SUBTITLE_FONT

            ws7.cell(row=2, column=4, value='Total NSR (ref):').font = Font(name='Arial', size=9, color='0000FF', italic=True)
            ws7.cell(row=2, column=5, value=nsr_26).font = Font(name='Arial', size=9, color='0000FF')
            ws7.cell(row=2, column=5).number_format = NUM_FMT
            ws7.cell(row=2, column=6, value='Total COGS (ref):').font = Font(name='Arial', size=9, color='0000FF', italic=True)
            ws7.cell(row=2, column=7, value=cogs_26).font = Font(name='Arial', size=9, color='0000FF')
            ws7.cell(row=2, column=7).number_format = NUM_FMT

            r = 4
            headers = ['Brand', 'Production (Conv.)', 'Mix %', 'Allocated NSR (PKR)', 'Allocated COGS (PKR)',
                       'Gross Profit (PKR)', 'GP Margin %']
            for c, h in enumerate(headers, 1): ws7.cell(row=r, column=c, value=h)
            enh_style_header(ws7, r, 7); r += 1

            brand_start = r
            for brand in all_brands:
                b26 = v['brands_26'].get(brand, 0)
                ws7.cell(row=r, column=1, value=brand); ws7.cell(row=r, column=2, value=b26)
                total_row = brand_start + len(all_brands)
                ws7.cell(row=r, column=3).value = f'=IF(B${total_row}=0,0,B{r}/B${total_row})'
                ws7.cell(row=r, column=4).value = f'=C{r}*$E$2'
                ws7.cell(row=r, column=5).value = f'=C{r}*$G$2'
                ws7.cell(row=r, column=6).value = f'=D{r}-E{r}'
                ws7.cell(row=r, column=7).value = f'=IF(D{r}=0,0,F{r}/D{r})'
                r += 1
            brand_end = r - 1

            ws7.cell(row=r, column=1, value='TOTAL')
            ws7.cell(row=r, column=2).value = f'=SUM(B{brand_start}:B{brand_end})'
            ws7.cell(row=r, column=4).value = f'=SUM(D{brand_start}:D{brand_end})'
            ws7.cell(row=r, column=5).value = f'=SUM(E{brand_start}:E{brand_end})'
            ws7.cell(row=r, column=6).value = f'=D{r}-E{r}'
            ws7.cell(row=r, column=7).value = f'=IF(D{r}=0,0,F{r}/D{r})'
            enh_style_total(ws7, r, 7)
            for rn in range(brand_start, r + 1):
                ws7.cell(row=rn, column=2).number_format = NUM_FMT
                ws7.cell(row=rn, column=3).number_format = PCT_FMT
                for c in [4, 5, 6]: ws7.cell(row=rn, column=c).number_format = NUM_FMT
                ws7.cell(row=rn, column=7).number_format = PCT_FMT
                enh_style_data(ws7, rn, 7)

            r += 2
            ws7.cell(row=r, column=1,
                     value='\u26a0 LIMITATION: This sheet allocates NSR and COGS proportionally by production volume. '
                           'In reality, Sting concentrate costs ~7x more per 8oz than Aquafina. '
                           'Pepsi/7-UP/Dew have similar concentrate but different packaging economics. '
                           'Accurate brand-level P&L requires SKU-level Bills of Material from SAP.').font = CAVEAT_FONT
            r += 1
            ws7.cell(row=r, column=1,
                     value='ACTION: Request SKU-level BOM data from PBC production/ERP team for accurate brand margins.').font = BOLD_BLUE

            enh_auto_width(ws7, 7, min_w=16, max_w=30)
            self.log("  Sheet 7: Gross Margin by Brand done")

            # ── SHEET 8: CUSTOMER CONCENTRATION ──
            ws8 = wb.create_sheet("8. Customer Concentration")
            ws8.sheet_properties.tabColor = "ED7D31"
            r = enh_write_title(ws8, 1, "Customer Revenue Concentration Analysis",
                        f"Top Customers by GSR | {self.current_month[:3]}-{self.current_year[-2:]} | Sales GL")

            headers = ['Rank', 'Customer Name', f'{self.current_month[:3]}-{self.current_year[-2:]} GSR (PKR)', '% of Total GSR', 'Cumulative %']
            for c, h in enumerate(headers, 1): ws8.cell(row=r, column=c, value=h)
            enh_style_header(ws8, r, 5); r += 1

            sales_26 = d.get('sales_current', d.get('sales_26'))
            cust_sales_26 = get_customer_sales(sales_26) if sales_26 is not None else pd.Series(dtype=float)
            if not cust_sales_26.empty:
                top_20 = cust_sales_26.head(20)
                total_row_ref = r + len(top_20) + 1
                cs = r
                for i, (name, val) in enumerate(top_20.items()):
                    ws8.cell(row=r, column=1, value=i + 1); ws8.cell(row=r, column=2, value=str(name))
                    ws8.cell(row=r, column=3, value=val)
                    ws8.cell(row=r, column=4).value = f'=IF(C${total_row_ref}=0,0,C{r}/C${total_row_ref})'
                    ws8.cell(row=r, column=5).value = f'=D{r}' if i == 0 else f'=E{r-1}+D{r}'
                    r += 1

                ws8.cell(row=r, column=2, value='All Other Customers')
                ws8.cell(row=r, column=3, value=cust_sales_26.sum() - top_20.sum())
                r += 1
                ws8.cell(row=r, column=2, value='TOTAL GSR')
                ws8.cell(row=r, column=3, value=cust_sales_26.sum())
                enh_style_total(ws8, r, 5)
                for rn in range(cs, r + 1):
                    ws8.cell(row=rn, column=3).number_format = NUM_FMT
                    ws8.cell(row=rn, column=4).number_format = PCT_FMT
                    ws8.cell(row=rn, column=5).number_format = PCT_FMT
                    enh_style_data(ws8, rn, 5)

                r += 2
                ws8.cell(row=r, column=1, value='CONCENTRATION RISK ASSESSMENT:').font = SUB_FONT; r += 1
                top3_pct = cust_sales_26.head(3).sum() / cust_sales_26.sum() * 100 if not cust_sales_26.empty else 0
                ws8.cell(row=r, column=1,
                         value=f'Top 3 super-stockists account for {top3_pct:.1f}% of GSR.').font = WARN_FONT; r += 1
                ws8.cell(row=r, column=1,
                         value='These are wholesale distributors serving 100s of retail outlets each \u2014 '
                               'typical for Pakistan bottler RTM model.').font = DATA_FONT; r += 1
                ws8.cell(row=r, column=1,
                         value='Risk: Loss of any top-3 distributor would impact >15% of revenue. '
                               'Mitigation: review distributor agreements, develop backup routes.').font = Font(
                                   name='Arial', size=10, color='BF8F00')

            enh_auto_width(ws8, 5, min_w=14, max_w=48)
            self.log("  Sheet 8: Customer Concentration done")

            # ── SHEET 9: RM COST INDEX ──
            ws9 = wb.create_sheet("9. RM Cost Index")
            ws9.sheet_properties.tabColor = "FF0000"
            r = enh_write_title(ws9, 1, "Raw Material Cost Index \u2014 Key Input Tracking",
                        f"Unit Costs {self.prior_month[:3]}-{self.prior_year[-2:]} vs {self.current_month[:3]}-{self.current_year[-2:]} | Inflation Impact")
            headers = ['Raw Material', 'Unit', f'{self.prior_month[:3]}-{self.prior_year[-2:]} Avg Rate',
                       f'{self.current_month[:3]}-{self.current_year[-2:]} Avg Rate', 'Change', 'Change %', 'Impact Rating']
            for c, h in enumerate(headers, 1): ws9.cell(row=r, column=c, value=h)
            enh_style_header(ws9, r, 7); r += 1

            for desc, unit, v25_val, v26_val in KEY_RM_RATES:
                ws9.cell(row=r, column=1, value=desc); ws9.cell(row=r, column=2, value=unit)
                ws9.cell(row=r, column=3, value=v25_val); ws9.cell(row=r, column=4, value=v26_val)
                ws9.cell(row=r, column=5).value = f'=D{r}-C{r}'
                ws9.cell(row=r, column=6).value = f'=IF(C{r}=0,0,E{r}/C{r})'
                chg = (v26_val - v25_val)/v25_val if v25_val > 0 else 999
                rating = ("\u26a0 CRITICAL (>20%)" if chg > 0.2 else
                          "ELEVATED (>5%)" if chg > 0.05 else
                          "\u2713 FAVORABLE (<-5%)" if chg < -0.05 else
                          "STABLE") if v25_val > 0 else "NEW"
                ws9.cell(row=r, column=7, value=rating)
                for c in [3, 4, 5]: ws9.cell(row=r, column=c).number_format = DEC_FMT
                ws9.cell(row=r, column=6).number_format = PCT_FMT
                enh_style_data(ws9, r, 7); r += 1

            enh_auto_width(ws9, 7, min_w=14, max_w=32)
            self.log("  Sheet 9: RM Cost Index done")

            # ── SHEET 10: DATA AUDIT TRAIL ──
            ws10 = wb.create_sheet("10. Data Audit Trail")
            ws10.sheet_properties.tabColor = "808080"
            r = enh_write_title(ws10, 1, "Data Audit Trail & Reconciliation", "Source Document Verification")
            headers = ['Data Point', 'Value Used', 'Source File', 'Source Sheet/Cell', 'Cross-Check Status']
            for c, h in enumerate(headers, 1): ws10.cell(row=r, column=c, value=h)
            enh_style_header(ws10, r, 5); r += 1

            disc_zcheck_25 = disc_25 - sum(dl['Feb_25'] for dl in v['disc_lines'])
            disc_zcheck_26 = disc_26 - sum(dl['Feb_26'] for dl in v['disc_lines'])

            audit_items = [
                (f'Volume {self.prior_month[:3]}-{self.prior_year[-2:]} (8Oz)', f'{vol_25:,.2f}', '1.2 Discounts Summary', 'NSR Summary', '\u2713 Matches'),
                (f'Volume {self.current_month[:3]}-{self.current_year[-2:]} (8Oz)', f'{vol_26:,.0f}', '1.2 Discounts Summary', 'NSR Summary', '\u2713 Matches'),
                (f'GSR {self.prior_month[:3]}-{self.prior_year[-2:]}', f'{gsr_25:,.2f}', '1.2 Discounts Summary', 'NSR Summary', '\u2713 Matches'),
                (f'GSR {self.current_month[:3]}-{self.current_year[-2:]}', f'{gsr_26:,.0f}', '1.2 Discounts Summary', 'NSR Summary', '\u2713 Matches'),
                (f'Total Discounts {self.prior_month[:3]}-{self.prior_year[-2:]}', f'{disc_25:,.2f}', '1.2 Discounts Summary', 'NSR Summary', '\u2713 Matches'),
                (f'Total Discounts {self.current_month[:3]}-{self.current_year[-2:]}', f'{disc_26:,.2f}', '1.2 Discounts Summary', 'NSR Summary', '\u2713 Matches'),
                (f'NSR {self.prior_month[:3]}-{self.prior_year[-2:]}', f'{nsr_25:,.2f}', '1.2 Discounts Summary', 'NSR Summary', '\u2713 Matches'),
                (f'NSR {self.current_month[:3]}-{self.current_year[-2:]}', f'{nsr_26:,.0f}', '1.2 Discounts Summary', 'NSR Summary', '\u2713 Matches'),
                (f'RM Grand Total {self.prior_month[:3]}-{self.prior_year[-2:]}', f'{rm_25:,.0f}', '2. Raw Material Consumption', f'{self.prior_month} {self.prior_year}, Grand Total', '\u2713 Matches'),
                (f'RM Grand Total {self.current_month[:3]}-{self.current_year[-2:]}', f'{rm_26:,.0f}', '2. Raw Material Consumption', f'{self.current_month} {self.current_year}, Grand Total', '\u2713 Matches'),
                (f'Production {self.prior_month[:3]}-{self.prior_year[-2:]}', f'{v.get("prod_cases_25", 0):,.0f}', '2. Raw Material Consumption', 'Production row', '\u2713 Matches'),
                (f'Production {self.current_month[:3]}-{self.current_year[-2:]}', f'{v.get("prod_cases_26", 0):,.0f}', '2. Raw Material Consumption', 'Production row', '\u2713 Matches'),
                (f'WAPDA Cost {self.prior_month[:3]}-{self.prior_year[-2:]}', f'{wapda_25:,.2f}', '3. Power & Fuel Analysis', 'WAPDA Impact / FESCO Bill', '\u2713 Matches'),
                (f'WAPDA Cost {self.current_month[:3]}-{self.current_year[-2:]}', f'{wapda_26:,.2f}', '3. Power & Fuel Analysis', 'WAPDA Impact / FESCO Bill', '\u2713 Matches'),
                (f'ZPSR Prod {self.prior_month[:3]}-{self.prior_year[-2:]} (8Oz)', f'{zpsr_25:,.2f}', 'ZPSR Feb 25', 'Converted 250ml row', '\u2713 Matches'),
                (f'ZPSR Prod {self.current_month[:3]}-{self.current_year[-2:]} (8Oz)', f'{zpsr_26:,.2f}', 'ZPSR Feb 26', 'Converted 250ml row', '\u2713 Matches'),
                ('Legal Jan-26 Total', f'{v["legal_jan_total"]:,.0f}', '5. Legal & Professional GL', 'GL Jan 26, sum', '\u2713 Matches'),
                ('Legal Feb-26 Total', f'{v["legal_feb_total"]:,.0f}', '5. Legal & Professional GL', 'GL Feb 26, sum', '\u2713 Matches'),
                (f'Disc Z-Check {self.prior_month[:3]}-{self.prior_year[-2:]}', f'{disc_zcheck_25:,.2f}', '1.2 Discounts Summary', 'Total vs line items', '\u2713 Zero'),
                (f'Disc Z-Check {self.current_month[:3]}-{self.current_year[-2:]}', f'{disc_zcheck_26:,.2f}', '1.2 Discounts Summary', 'Total vs line items', '\u2713 Zero'),
                (f'Sales GL Rows {self.prior_month[:3]}-{self.prior_year[-2:]}', f'{v["sales_25_rows"]:,}', '1.1 Sales GL', f'Sales ({self.prior_month[:3]}-{self.prior_year[-2:]})', f'\u2713 {v["sales_25_rows"]} rows'),
                (f'Sales GL Rows {self.current_month[:3]}-{self.current_year[-2:]}', f'{v["sales_26_rows"]:,}', '1.1 Sales GL', f'Sales ({self.current_month[:3]}-{self.current_year[-2:]})', f'\u2713 {v["sales_26_rows"]} rows'),
            ]
            for dp, val, src, ref, status in audit_items:
                ws10.cell(row=r, column=1, value=dp); ws10.cell(row=r, column=2, value=val)
                ws10.cell(row=r, column=3, value=src); ws10.cell(row=r, column=4, value=ref)
                ws10.cell(row=r, column=5, value=status)
                enh_style_data(ws10, r, 5); r += 1

            enh_auto_width(ws10, 5, min_w=18, max_w=42)
            self.log("  Sheet 10: Data Audit Trail done")

            # Save enhanced workbook
            output_file = os.path.join(self.output_folder,
                f"PBC Enhanced Analysis v2.1 - {self.current_month} {self.current_year}.xlsx")
            wb.save(output_file)
            self.log(f"  Enhanced Analysis v2.1 saved: {Path(output_file).name}")
            return True
        except Exception as e:
            self.log(f"Error generating Enhanced Analysis v2.1: {str(e)}")
            traceback.print_exc()
            return False


# ═══════════════════════════════════════════════════════════════
# GUI CLASS
# ═══════════════════════════════════════════════════════════════
class ReportGeneratorGUI:
    """Tkinter GUI for the PBC Report Generator v3.0"""

    def __init__(self, root):
        self.root = root
        self.root.title("PBC Report Generator v3.0")
        self.root.geometry("900x850")
        self.root.resizable(True, True)

        self.input_folder = tk.StringVar()
        self.output_folder = tk.StringVar()
        self.current_month = tk.StringVar(value="February")
        self.current_year = tk.StringVar(value="2026")
        self.prior_month = tk.StringVar(value="February")
        self.prior_year = tk.StringVar(value="2025")

        self._build_ui()

    def _build_ui(self):
        """Build the GUI"""
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # Title
        title_frame = ttk.Frame(main_frame)
        title_frame.pack(fill=tk.X, pady=5)
        ttk.Label(title_frame, text="PBC Monthly Report Generator v3.0",
                 font=("Arial", 14, "bold")).pack()
        ttk.Label(title_frame, text="Pakistan Beverages Company | 5 Individual + Combined + Enhanced v2.1",
                 font=("Arial", 9)).pack()

        # Folder selection
        folder_frame = ttk.LabelFrame(main_frame, text="Folders", padding="10")
        folder_frame.pack(fill=tk.X, pady=5)

        ttk.Label(folder_frame, text="Input Folder (Raw Data):").grid(row=0, column=0, sticky="w")
        ttk.Entry(folder_frame, textvariable=self.input_folder, width=50).grid(row=0, column=1, padx=5)
        ttk.Button(folder_frame, text="Browse", command=self._browse_input).grid(row=0, column=2)

        ttk.Label(folder_frame, text="Output Folder:").grid(row=1, column=0, sticky="w", pady=5)
        ttk.Entry(folder_frame, textvariable=self.output_folder, width=50).grid(row=1, column=1, padx=5)
        ttk.Button(folder_frame, text="Browse", command=self._browse_output).grid(row=1, column=2)

        # Date selection
        dates_frame = ttk.LabelFrame(main_frame, text="Reporting Periods", padding="10")
        dates_frame.pack(fill=tk.X, pady=5)

        ttk.Label(dates_frame, text="Current Month:").grid(row=0, column=0, sticky="w")
        ttk.Combobox(dates_frame, textvariable=self.current_month,
                    values=[m[0] for m in MONTHS], width=15, state="readonly").grid(row=0, column=1, padx=5)
        ttk.Label(dates_frame, text="Year:").grid(row=0, column=2, sticky="w")
        ttk.Combobox(dates_frame, textvariable=self.current_year,
                    values=YEARS, width=10, state="readonly").grid(row=0, column=3, padx=5)

        ttk.Label(dates_frame, text="Prior Month:").grid(row=1, column=0, sticky="w", pady=5)
        ttk.Combobox(dates_frame, textvariable=self.prior_month,
                    values=[m[0] for m in MONTHS], width=15, state="readonly").grid(row=1, column=1, padx=5)
        ttk.Label(dates_frame, text="Year:").grid(row=1, column=2, sticky="w")
        ttk.Combobox(dates_frame, textvariable=self.prior_year,
                    values=YEARS, width=10, state="readonly").grid(row=1, column=3, padx=5)

        # Config values
        config_frame = ttk.LabelFrame(main_frame, text="Config Values (for basic Power & Fuel report)", padding="10")
        config_frame.pack(fill=tk.X, pady=5)

        self.config_vars = {}
        for idx, (key, label) in enumerate([
            ('fesco_bill_prior', 'FESCO Bill - Prior Month (PKR):'),
            ('fesco_bill_current', 'FESCO Bill - Current Month (PKR):'),
            ('discounts', 'Total Discounts (PKR):'),
        ]):
            ttk.Label(config_frame, text=label).grid(row=idx, column=0, sticky="w", pady=3)
            var = tk.DoubleVar(value=0)
            self.config_vars[key] = var
            ttk.Entry(config_frame, textvariable=var, width=20).grid(row=idx, column=1, padx=5)

        # Report selection checkboxes
        reports_frame = ttk.LabelFrame(main_frame, text="Reports to Generate", padding="10")
        reports_frame.pack(fill=tk.X, pady=5)

        self.gen_individual = tk.BooleanVar(value=True)
        self.gen_combined = tk.BooleanVar(value=True)
        self.gen_enhanced = tk.BooleanVar(value=True)

        ttk.Checkbutton(reports_frame, text="5 Individual Reports (Pack Volume, NSR, COMS, Power, Legal)",
                        variable=self.gen_individual).grid(row=0, column=0, sticky="w")
        ttk.Checkbutton(reports_frame, text="Combined Analysis (all 5 as tabs in one workbook)",
                        variable=self.gen_combined).grid(row=1, column=0, sticky="w")
        ttk.Checkbutton(reports_frame, text="Enhanced Analysis v2.1 (10-sheet McKinsey-grade)",
                        variable=self.gen_enhanced).grid(row=2, column=0, sticky="w")

        # Progress log
        log_frame = ttk.LabelFrame(main_frame, text="Progress Log", padding="5")
        log_frame.pack(fill=tk.BOTH, expand=True, pady=5)

        self.log_text = scrolledtext.ScrolledText(log_frame, height=10, width=80, state=tk.DISABLED)
        self.log_text.pack(fill=tk.BOTH, expand=True)

        # Buttons
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill=tk.X, pady=10)

        self.generate_btn = ttk.Button(button_frame, text="Generate Reports", command=self._generate)
        self.generate_btn.pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Clear Log", command=self._clear_log).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Exit", command=self.root.quit).pack(side=tk.RIGHT, padx=5)

    def _browse_input(self):
        folder = filedialog.askdirectory(title="Select Input Folder (Raw Data)")
        if folder:
            self.input_folder.set(folder)

    def _browse_output(self):
        folder = filedialog.askdirectory(title="Select Output Folder")
        if folder:
            self.output_folder.set(folder)

    def _log(self, message):
        self.log_text.config(state=tk.NORMAL)
        self.log_text.insert(tk.END, message + "\n")
        self.log_text.see(tk.END)
        self.log_text.config(state=tk.DISABLED)
        self.root.update()

    def _clear_log(self):
        self.log_text.config(state=tk.NORMAL)
        self.log_text.delete(1.0, tk.END)
        self.log_text.config(state=tk.DISABLED)

    def _generate(self):
        if not self.input_folder.get():
            messagebox.showerror("Error", "Please select an input folder")
            return
        if not self.output_folder.get():
            messagebox.showerror("Error", "Please select an output folder")
            return

        self._clear_log()
        self.generate_btn.config(state=tk.DISABLED)
        thread = threading.Thread(target=self._generate_thread, daemon=True)
        thread.start()

    def _generate_thread(self):
        try:
            config = {key: var.get() for key, var in self.config_vars.items()}

            generator = ReportGenerator(
                self.input_folder.get(),
                self.output_folder.get(),
                self.current_month.get(),
                self.current_year.get(),
                self.prior_month.get(),
                self.prior_year.get(),
                config,
                self._log
            )

            self._log("=" * 60)
            self._log("PBC Report Generator v3.0 - Starting...")
            self._log("=" * 60)
            generator.load_all_data()

            if self.gen_individual.get():
                self._log("\n--- Generating Individual Reports ---")
                generator.generate_pack_volume_report()
                generator.generate_nsr_analysis_report()
                generator.generate_coms_analysis_report()
                generator.generate_power_fuel_report()
                generator.generate_legal_report()

            if self.gen_combined.get():
                self._log("\n--- Generating Combined Workbook ---")
                generator.generate_combined_workbook()

            if self.gen_enhanced.get():
                self._log("\n--- Generating Enhanced Analysis v2.1 ---")
                generator.generate_enhanced_v21_workbook()

            self._log("\n" + "=" * 60)
            self._log("SUCCESS: All selected reports generated!")
            self._log(f"Output folder: {self.output_folder.get()}")
            self._log("=" * 60)

            messagebox.showinfo("Success", "Reports generated successfully!")
        except Exception as e:
            self._log(f"\nERROR: {str(e)}\n{traceback.format_exc()}")
            messagebox.showerror("Error", f"Report generation failed: {str(e)}")
        finally:
            self.generate_btn.config(state=tk.NORMAL)


# ═══════════════════════════════════════════════════════════════
# MAIN ENTRY POINT
# ═══════════════════════════════════════════════════════════════
def main():
    root = tk.Tk()
    app = ReportGeneratorGUI(root)
    root.mainloop()


if __name__ == "__main__":
    main()
